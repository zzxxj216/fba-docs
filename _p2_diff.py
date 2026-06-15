# -*- coding: utf-8 -*-
"""P2 验收：本系统生成文件 vs 跨运通 RPA 历史成品 逐单元格 diff。

用法:  python _p2_diff.py            # 控制台摘要 + 返回结构
分类:  MATCH / 日期类差异 / 数据缺口 / RPA bug / 我方错误
比对口径:
- 两边任一非空的单元格全部比对（含公式串）；
- 数值与数值字符串按数值相等判 MATCH（RPA 经 Excel COM 写入存在
  字符串/数字类型抖动，如 托书 B16='12'、I18='2'）；
- 日期对象与 YYYY-MM-DD / YYYY/MM/DD 字符串同日判 MATCH。
"""

import datetime as _dt
import os
import re

from openpyxl import load_workbook

SRC = r"C:\Users\zane\Downloads\Serenorch-US-6.5"
OUT = r"D:\amazon\output\Serenorch-US-0612"
FCS = ["ABE8", "GYR2", "IND9", "LAS1", "LGB8"]

# ---- 文件配对：(标签, 我方文件, RPA 成品) ----
PAIRS = []
for fc in FCS:
    PAIRS.append((f"托书-{fc}",
                  rf"{OUT}\{fc}\Serenorch-托书-{fc}-0605.xlsx",
                  rf"{SRC}\Serenorch-US-6.5-托书\Serenorch-托书-{fc}-0605.xlsx"))
for fc in FCS:
    PAIRS.append((f"报关-{fc}",
                  rf"{OUT}\{fc}\Serenorch-US-{fc}-报关资料-0605.xlsx",
                  rf"{SRC}\Serenorch-US-6.5-报关\Serenorch-US-{fc}-报关资料-0605.xlsx"))
for fc in FCS:
    PAIRS.append((f"投保-{fc}",
                  rf"{OUT}\{fc}\Serenorch-投保-06-05-{fc}.xlsx",
                  rf"{SRC}\Serenorch-US-6.5-投保\Serenorch-投保-06-05-{fc}.xlsx"))
PAIRS.append(("采购合同",
              rf"{OUT}\Serenorch - 采购合同-2026-06-05.xlsx",
              rf"{SRC}\Serenorch-US-6.5-采购合同\Serenorch - 采购合同-2026-06-05.xlsx"))
PAIRS.append(("9810",
              rf"{OUT}\Serenorch - US - 报关9810-20260612.xlsx",
              rf"{SRC}\Serenorch-US-6.5-9810\Serenorch - US - 报关9810-20260529 .xlsx"))

# ---- 定性规则 ----
# (文件标签前缀, sheet, 单元格正则) → (分类, 说明)
RULES = [
    # RPA bug：箱数误写进托书"地址库"sheet B16（影刀空 sheet 名事故，CLAUDE.md 约定 2 的出处）
    ("托书", "地址库", r"^B16$", "RPA bug",
     "RPA 未指定 sheet 名，把货件总箱数写进了地址库!B16，覆盖地址简称数据"),
    # RPA bug：报关单申报要素行（D21）RPA 对合并单元格写入未生效，成品缺失；本系统按映射补全
    ("报关", "报关单", r"^D21$", "RPA bug",
     "RPA 写 D21(合并区 D21:F22) 申报要素未生效导致成品缺失；本系统补全 item.declare_full"),
    # RPA bug：投保"起运日期"U3 RPA 写入未生效，成品为空；本系统按模板要求填基准周五
    ("投保", "批量导入投保数据", r"^U3$", "RPA bug",
     "RPA 起运日期写入未生效成品为空；本系统填基准周五(YYYY/MM/DD)"),
    # 日期类：9810 报送时间 = 生成时刻（RPA 跑批 2026-05-29，本次 2026-06-12）
    ("9810", "订单导入模板", r"^B\d+$", "日期类差异",
     "报送时间=生成时刻，RPA 历史跑批为 20260529，本次生成 20260612"),
    # 数据缺口：商品列表-报关 自 RPA 跑批后被更新过（源数据版本漂移）
    ("托书", "模板", r"^F18$", "数据缺口",
     "箱规高：商品列表现值 77cm(=30.31in)，RPA 跑批时为 44cm(=17.32in)"),
    ("托书", "模板", r"^G18$", "数据缺口",
     "中文报关名：现值'一次性面巾纸'，RPA 跑批时商品列表为'一次性面巾纸 EF纹 1PC'"),
    ("托书", "模板", r"^H18$", "数据缺口",
     "英文报关名：现值'disposable face towel 1pc 50count'，跑批时为'Disposable face towels 50pc EF'"),
    ("托书", "模板", r"^L18$", "数据缺口",
     "海关编码：现值 6302930010，跑批时为 6302930090"),
    ("托书", "模板", r"^M18$", "数据缺口",
     "分类(用途)：现值'未分类'，跑批时为'一次性毛巾'"),
    ("报关", "报关单", r"^B20$", "数据缺口",
     "海关编码：现值 6302930010，跑批时为 6302930090"),
    ("报关", "合同", r"^A18$", "数据缺口",
     "合同品名：现值'Serenorch-FT-1（一次性面巾纸）'，跑批时为'一次性面巾纸（Serenorch-FT-1）'(括号顺序互换)"),
    ("报关", "发票", r"^C6$", "数据缺口",
     "发票品名：同上，商品列表'货物名称(发票箱单)'列括号顺序已互换"),
    ("报关", "箱单", r"^B8$", "数据缺口",
     "箱单品名：同上"),
]

_NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")


def norm(v):
    """归一化：数值字符串→float、日期→date、其余 strip 后原样。"""
    if isinstance(v, _dt.datetime):
        return ("date", v.date()) if (v.hour, v.minute, v.second) == (0, 0, 0) \
            else ("num", v.timestamp())
    if isinstance(v, _dt.date):
        return ("date", v)
    if isinstance(v, bool):
        return ("num", float(v))
    if isinstance(v, (int, float)):
        return ("num", float(v))
    if isinstance(v, str):
        s = v.strip()
        if _NUM_RE.match(s):
            return ("num", float(s))
        m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
        if m:
            return ("date", _dt.date(*map(int, m.groups())))
        return ("str", s)
    return ("str", str(v))


def equal(a, b):
    na, nb = norm(a), norm(b)
    if na == nb:
        return True
    if na[0] == nb[0] == "num":
        return abs(na[1] - nb[1]) < 1e-9
    return False


def classify(label, sheet, addr, ours, theirs):
    for pfx, sh, pat, cat, note in RULES:
        if label.startswith(pfx) and sheet == sh and re.match(pat, addr):
            return cat, note
    return "我方错误", "未定性差异，需修映射/引擎"


def cells(path):
    wb = load_workbook(path)
    out = {}
    for sn in wb.sheetnames:
        d = {}
        for row in wb[sn].iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip() != "":
                    d[c.coordinate] = c.value
        out[sn] = d
    wb.close()
    return out


def diff_pair(label, ours_path, rpa_path):
    rows = []
    stats = {"total": 0, "MATCH": 0, "日期类差异": 0, "数据缺口": 0,
             "RPA bug": 0, "我方错误": 0}
    if not os.path.exists(ours_path):
        rows.append((label, "-", "-", "(文件缺失)", "-", "我方错误", "我方文件未生成"))
        stats["我方错误"] += 1
        return stats, rows
    a, b = cells(ours_path), cells(rpa_path)
    for sn in sorted(set(a) | set(b)):
        ca, cb = a.get(sn, {}), b.get(sn, {})
        for addr in sorted(set(ca) | set(cb),
                           key=lambda k: (int(re.sub(r"\D", "", k) or 0), k)):
            va, vb = ca.get(addr), cb.get(addr)
            stats["total"] += 1
            if equal(va, vb):
                stats["MATCH"] += 1
                continue
            cat, note = classify(label, sn, addr, va, vb)
            stats[cat] += 1
            rows.append((label, sn, addr, repr(va), repr(vb), cat, note))
    return stats, rows


def main():
    all_rows = []
    summary = []
    for label, ours, rpa in PAIRS:
        stats, rows = diff_pair(label, ours, rpa)
        summary.append((label, stats))
        all_rows.extend(rows)
        flag = "OK " if stats["我方错误"] == 0 else "FAIL"
        print(f"[{flag}] {label}: 比对 {stats['total']} 格 | MATCH {stats['MATCH']} "
              f"| 日期 {stats['日期类差异']} | 数据缺口 {stats['数据缺口']} "
              f"| RPAbug {stats['RPA bug']} | 我方错误 {stats['我方错误']}")
    print()
    for r in all_rows:
        if r[5] == "我方错误":
            print("  ERR", r[0], r[1], r[2], "ours=", r[3][:80], "| rpa=", r[4][:80])
    return summary, all_rows


if __name__ == "__main__":
    main()

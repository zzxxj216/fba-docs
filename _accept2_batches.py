# -*- coding: utf-8 -*-
"""验收2 批次导入+数据回填（幂等）：
- Serenorch-US-6.19（跨运通复验票，旧版发货单导出）
- HUHOLE-US-6.5 跨运通（新版 FBA货件导出）
- BFPeaky-US-6.5 跨运通（新版 FBA货件导出）
- HUHOLE-US-6.5 盈和（无源发货单——按补仓计划/货代标/成品托书回填重建，
  回填项在 ACCEPTANCE_REPORT2.md 中说明）
"""
import sys

sys.path.insert(0, r"D:\amazon")

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import Batch, Brand, Forwarder, Product, Shipment, ShipmentItem
from app.services import import_service

CASE = r"C:\Users\zane\Downloads\HUBFSE自动化"
db = SessionLocal()
FWD_K = db.query(Forwarder).filter(Forwarder.name.like("%跨运通%")).first().id
FWD_Y = db.query(Forwarder).filter(Forwarder.name.like("%盈和%")).first().id


def finish_batch(batch_id, name, base, contract, fwd_id, order_nos):
    b = db.get(Batch, batch_id)
    b.name = name
    b.base_date = base
    b.contract_date = contract
    for sp in b.shipments:
        sp.forwarder_id = fwd_id
        if sp.fc_code in order_nos:
            sp.forwarder_order_no = order_nos[sp.fc_code]
    db.commit()
    return b


# ================================================= 1. Serenorch-US-6.19
print("== Serenorch-US-6.19 ==")
SE = CASE + r"\Serenorch-US-6.19"
r = import_service.import_products(db, SE + r"\商品列表-报关xlsx.xlsx")
print("  products:", r)
r = import_service.import_shipping_order(
    db, SE + r"\发货单-20260611-1781170740211.xlsx",
    filename="发货单-20260611-1781170740211.xlsx", batch_name="Serenorch-US-0619")
print("  batch:", r["batch_id"], "passed:", r["report"]["passed"])
se_batch_id = r["batch_id"]
# 货代标 XX266-270；合同日期=RPA口径(基准周五−30天=05-20，成品 SE-ABE8-2026-05-20)
se = finish_batch(se_batch_id, "Serenorch-US-0619", "2026-06-19", "2026-05-20",
                  FWD_K, {"ABE8": "XX266", "GYR2": "XX267", "IND9": "XX268",
                          "MDW2": "XX269", "SCK4": "XX270"})
# 地址回填：托书模板「地址库」按 FC（与 P2 同手段）
wb = load_workbook(r"D:\amazon\templates_store\20260612151726_托书-空白模板.xlsx",
                   read_only=True, data_only=True)
addr = {}
for row in wb["地址库"].iter_rows(values_only=True):
    fc = str(row[0] or "").strip()
    if fc:
        addr[fc] = {"address_line1": row[7], "city": row[10],
                    "state": row[11], "postal_code": row[13]}
wb.close()
for sp in se.shipments:
    a = addr.get(sp.fc_code)
    if a:
        for k, v in a.items():
            if not getattr(sp, k):
                setattr(sp, k, str(v or ""))
# 产品图片回填：成品托书 O(msku)/Q(图片) 列（赛狐发货单导出含图片列，离线文件未存）
import glob
for f in glob.glob(SE + r"\Serenorch-US-6.19-托书\*.xlsx"):
    if "~$" in f:
        continue
    twb = load_workbook(f, read_only=True, data_only=True)
    ws = twb["模板"]
    for row in ws.iter_rows(min_row=18, values_only=True):
        msku, img = row[14], row[16]          # O / Q
        if msku and img:
            p = db.query(Product).filter(Product.sku == str(msku).strip()).first()
            if p is not None and not (p.image_url or "").strip():
                p.image_url = str(img).strip()
    twb.close()
db.commit()
print("  SE batch ready:", se.id, se.name, se.base_date, se.contract_date)

# ================================================= 2. HUHOLE-US-6.5 跨运通
print("== HUHOLE-US-6.5 跨运通 ==")
HU = CASE + r"\HUHOLE-US-6.5"
r = import_service.import_shipping_order(
    db, HU + r"\FBA货件-920336340792705024.xlsx",
    filename="FBA货件-920336340792705024.xlsx", batch_name="HUHOLE-US-0605K")
print("  batch:", r["batch_id"], "passed:", r["report"]["passed"])
for e in r["report"]["errors"]:
    print("   ", e["level"], e["msg"][:80])
hu_k = finish_batch(r["batch_id"], "HUHOLE-US-0605K", "2026-06-05", "2026-05-05",
                    FWD_K, {"GYR2": "XX209", "IND9": "XX210", "LGB8": "XX211",
                            "MQJ1": "XX212", "VGT2": "XX213"})
print("  HU-K ready:", hu_k.id)

# ================================================= 3. BFPeaky-US-6.5
print("== BFPeaky-US-6.5 ==")
BF = CASE + r"\BFPeaky-US-6.5"
r = import_service.import_shipping_order(
    db, BF + r"\FBA货件-920329846391103488.xlsx",
    filename="FBA货件-920329846391103488.xlsx", batch_name="BFPeaky-US-0605")
print("  batch:", r["batch_id"], "passed:", r["report"]["passed"])
for e in r["report"]["errors"]:
    print("   ", e["level"], e["msg"][:80])
bf_b = finish_batch(r["batch_id"], "BFPeaky-US-0605", "2026-06-05", "2026-05-05",
                    FWD_K, {"GYR3": "XX204", "IND9": "XX205", "RFD2": "XX206",
                            "RMN3": "XX207", "SBD1": "XX208"})
print("  BF ready:", bf_b.id)

# ================================================= 4. HUHOLE-US-6.5 盈和（重建）
print("== HUHOLE-US-6.5 盈和 ==")
PLAN_ID = "offline-yinghe-hu0605"
b = db.query(Batch).filter(Batch.inbound_plan_id == PLAN_ID).first()
if b is None:
    b = Batch(inbound_plan_id=PLAN_ID)
    db.add(b)
brand = db.query(Brand).filter(Brand.name == "HUHOLE").first()
b.name = "HUHOLE-US-0605Y"
b.shop_name = "HUHOLE-US"
b.marketplace = b.country = "US"
b.brand_id = brand.id
b.company_id = brand.company_id
b.factory_id = brand.factory_id
b.base_date = "2026-06-05"
b.contract_date = "2026-05-05"
db.flush()

# 货件骨架：FC/FBA号 来自货代标 PDF 文件名；RefID/逐 SKU 箱数拆分 来自成品托书
# （源发货单缺失的数据缺口回填，见报告）；产品主数据来自盈和补仓计划。
YINGHE = {
    # fc: (fba, ref, [(msku, box_count), ...])  箱数/顺序按成品托书
    "ABE8": ("FBA19FD67STG", "4I5CFOJR",
             [("Huhole-GS-Black-12", 5), ("Huhole-PS-Black-12", 8)]),
    "GYR2": ("FBA19FD668Z6", None, None),
    "IND9": ("FBA19FD16TBG", None, None),
    "LAS1": ("FBA19FD210XH", None, None),
    "SBD1": ("FBA19FD357RL", None, None),
}
# 从成品托书读 ref + 行项（避免手抄错；属于"按成品回填缺失源数据"）
import os
tdir = HU + r"\HUHOLE-US-6.5\HUHOLE-US-6.5盈和\HUHOLE-US-6.5托书1"
for fc in YINGHE:
    f = os.path.join(tdir, f"HUHOLE-普船-{fc}-06.05-托书.xlsx")
    twb = load_workbook(f, read_only=True, data_only=True)
    ws = twb["Sheet1"]
    grid = {c.coordinate: c.value for row in ws.iter_rows() for c in row
            if c.value is not None}
    ref = grid.get("M4")
    items = []
    rr = 23
    while grid.get(f"J{rr}"):
        items.append((str(grid[f"J{rr}"]), int(grid[f"B{rr}"])))
        rr += 1
    YINGHE[fc] = (YINGHE[fc][0], str(ref), items)
    twb.close()

existing = {sp.fc_code: sp for sp in b.shipments}
for fc, (fba, ref, items) in YINGHE.items():
    sp = existing.get(fc)
    if sp is None:
        sp = Shipment(batch_id=b.id, fc_code=fc)
        db.add(sp)
    sp.amazon_shipment_id = fba
    sp.reference_id = ref
    sp.carton_num = sum(n for _, n in items)
    sp.country_code = "US"
    sp.forwarder_id = FWD_Y
    sp.remark = "offline;盈和(按补仓计划/货代标/成品托书回填)"
    db.flush()
    have = {it.msku: it for it in sp.items}
    for msku, boxes in items:
        it = have.get(msku)
        if it is None:
            it = ShipmentItem(shipment_id=sp.id, msku=msku)
            db.add(it)
        p = db.query(Product).filter(Product.sku == msku).first()
        it.product_id = p.id if p else None
        it.box_count = boxes
        it.qty_per_box = p.qty_per_box if p else None
        it.qty = boxes * (p.qty_per_box or 0) if p else None
        it.customs_unit_price = p.unit_price_default if p else None
        it.purchase_cost = p.purchase_cost_default if p else None
db.commit()
from app.services import validate_service
rep = validate_service.validate_batch(db, b.id)
print("  盈和批次:", b.id, "passed:", rep["passed"])
for e in rep["errors"]:
    print("   ", e["level"], e["msg"][:80])
db.close()
print("done")

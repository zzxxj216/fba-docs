# -*- coding: utf-8 -*-
"""验收2 配置脚本（幂等）：HUHOLE/BFPeaky 主数据 + 产品 + 新模板（盈和/跨运通新版）。

与 API 同一代码路径（直接调 service/ORM）；运行中的 8000 服务带 --reload，
本脚本只写库/写 templates_store，不影响服务进程。
"""
import json
import os
import shutil
import sys

sys.path.insert(0, r"D:\amazon")

from openpyxl import load_workbook

from app.database import SessionLocal, TEMPLATE_STORE
from app.models import Brand, Company, Factory, Forwarder, Product, RuleConfig, Template

CASE = r"C:\Users\zane\Downloads\HUBFSE自动化"

db = SessionLocal()


def upsert_company(name_cn, **kw):
    c = db.query(Company).filter(Company.name_cn == name_cn).first()
    if c is None:
        c = Company(name_cn=name_cn)
        db.add(c)
    for k, v in kw.items():
        setattr(c, k, v)
    db.flush()
    return c


def upsert_factory(name, **kw):
    f = db.query(Factory).filter(Factory.name == name).first()
    if f is None:
        f = Factory(name=name)
        db.add(f)
    for k, v in kw.items():
        setattr(f, k, v)
    db.flush()
    return f


def upsert_brand(name, **kw):
    b = db.query(Brand).filter(Brand.name == name).first()
    if b is None:
        b = Brand(name=name)
        db.add(b)
    for k, v in kw.items():
        setattr(b, k, v)
    db.flush()
    return b


def upsert_rule(key, scope, value, label=""):
    r = (db.query(RuleConfig)
         .filter(RuleConfig.key == key, RuleConfig.scope == scope).first())
    if r is None:
        r = RuleConfig(key=key, scope=scope)
        db.add(r)
    r.value = value
    if label:
        r.label = label
    db.flush()
    return r


# ================================================================ 1. 主数据
print("== 主数据 ==")
# 公司：舟峰(HUHOLE 店铺主体/境内发货人)、保峰五金(BFPeaky 主体；同名工厂另立)
zhoufeng = upsert_company(
    "杭州舟峰科技有限公司", type="shop",
    name_en="Hangzhou Zhoufeng Technology Co., Ltd.",
    address_cn="浙江省杭州市萧山区临浦镇康家湖路7号1幢临浦智创园1-436",
    address_en="Building 1, Linpu Zhichuang Park, No. 7 Kangjiahu Road, "
               "Linpu Town, Xiaoshan District, Hangzhou City",
    uscc="91330114MA2KL9FL58", phone="17816874933",
    export_via_trade=True, insurance_factor=1.695,
    default_port="宁波", default_origin_place="杭州", active=True)
baofeng_co = upsert_company(
    "杭州保峰五金制造有限公司", type="shop",
    name_en="Hangzhou Baofeng Hardware Manufacturing Co., Ltd.",
    address_cn="杭州市萧山区义桥镇新坝村",
    address_en="Xinba Village, Yiqiao Town, Xiaoshan District, Hangzhou City",
    uscc="91330109793691699K", phone="13758258357",
    export_via_trade=True, insurance_factor=1.695,
    default_port="宁波", default_origin_place="杭州", active=True)
# 星盟(trade)：补地址/电话（BFPeaky 采购合同买方=星盟，模板预填该地址）
trade = db.query(Company).filter(Company.type == "trade").first()
if trade is not None:
    if not (trade.address_cn or "").strip():
        trade.address_cn = "浙江省杭州市萧山区城厢街道蓝橙智韵创新中心2幢-601室"
    if not (trade.phone or "").strip():
        trade.phone = "15158176946"

baofeng_factory = upsert_factory(
    "杭州保峰五金制造有限公司", abbr="BF",
    address="杭州市萧山区义桥镇新坝村", phone="13758258357", active=True)

hu = upsert_brand("HUHOLE", abbr2="HU",
                  factory_id=baofeng_factory.id, company_id=zhoufeng.id,
                  doc_no_rule_shipment="BY-{sa}{brand2}-{fc}-{date}",
                  doc_no_rule_purchase="{sa}{brand2}-{date}", active=True)
bf = upsert_brand("BFPeaky", abbr2="BF",
                  factory_id=baofeng_factory.id, company_id=baofeng_co.id,
                  doc_no_rule_shipment="BY-{sa}{brand2}-{fc}-{date}",
                  doc_no_rule_purchase="{sa}{brand2}-{date}", active=True)

upsert_rule("contract_price_factor", "global", "1.0", "采购合同单价系数(成本×系数)")
upsert_rule("contract_price_factor", f"company:{zhoufeng.id}", "1.695",
            "舟峰：合同单价=成本×1.5×1.13")
upsert_rule("contract_price_factor", f"company:{baofeng_co.id}", "1.695",
            "保峰：合同单价=成本×1.5×1.13")
db.commit()
print("  company 舟峰=%s 保峰=%s factory 保峰=%s brand HU=%s BF=%s" % (
    zhoufeng.id, baofeng_co.id, baofeng_factory.id, hu.id, bf.id))

# ================================================================ 2. 产品
print("== 产品 ==")


def import_plan_products(plan_path, price_path, price_sheet, declare,
                         material_en=None):
    """补仓计划(产品主数据) + 报关价格试算(当周报关价) → Product upsert。
    按补仓计划行序插入（product_id 序 = 计划行序 = RPA 采购合同行序）。"""
    wb = load_workbook(plan_path, data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr) if h}

    prices = {}
    if price_path and price_sheet:
        wb = load_workbook(price_path, data_only=True, read_only=True)
        for r in wb[price_sheet].iter_rows(values_only=True):
            if r and len(r) > 6 and r[2] and r[6] is not None \
                    and str(r[2]).strip() != "sku":
                try:
                    prices[str(r[2]).strip()] = float(r[6])
                except (TypeError, ValueError):
                    continue                      # 表头/汇总行

        wb.close()

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    n = 0
    for r in rows[1:]:
        sku = g(r, "AMAZON-SKU")
        if not sku:
            continue
        sku = str(sku).strip()
        p = db.query(Product).filter(Product.sku == sku).first()
        if p is None:
            p = Product(sku=sku)
            db.add(p)
        p.name_customs_cn = str(g(r, "中文品名") or "").strip()
        p.name_customs_en = str(g(r, "英文品名") or "").strip()
        hs = g(r, "海关编码")
        if hs:
            p.hs_code = str(hs).strip()
        p.material = str(g(r, "材质") or "").strip()
        if material_en is not None:
            p.material_en = material_en
        p.usage = str(g(r, "用途") or "").strip()
        p.brand_name = str(g(r, "品牌") or "").strip()
        if g(r, "每箱重量（kg )") is not None:
            p.box_weight_kg = float(g(r, "每箱重量（kg )"))
        if g(r, "每箱商品数") is not None:
            p.qty_per_box = int(g(r, "每箱商品数"))
        if g(r, "单品单价(¥)") is not None:
            p.purchase_cost_default = float(g(r, "单品单价(¥)"))
        if sku in prices:
            p.unit_price_default = prices[sku]
        if declare:
            p.declare_elements = declare
        n += 1
        p.sort_index = n          # 当周计划行序（采购合同聚合行序）
        db.flush()
    db.commit()
    return n


HU_DECLARE = "用途：车库挂钩|材质：铁|品牌：HUHOLE"
BF_DECLARE = "用途：车库挂钩|材质：铁|品牌：BFPeaky"

n = import_plan_products(
    CASE + r"\HUHOLE-US-6.5\补仓计划-US-6.5-跨运通.xlsx",
    CASE + r"\HUHOLE-US-6.5\HU2026报关价格试算.xlsx", "6.5HU-US", HU_DECLARE)
print("  HUHOLE 6.5 跨运通产品:", n)

# 盈和 2 个 SKU（无 6.5 盈和补仓计划——用 5.29 盈和计划做主数据，
# 中文品名按试算/成品口径去掉"（盈和）"后缀；英文材质来源=盈和托书成品口径）
n = import_plan_products(
    CASE + r"\HUHOLE-US-5.29\补仓计划-US-5.29-盈和.xlsx",
    CASE + r"\HUHOLE-US-5.29\HU2026报关价格试算.xlsx", "1.23HU-US-2", "",
    material_en="Wood + Steel")
for sku, name_cn, name_en in (
        ("Huhole-GS-Black-12", "黑色网版支架12PC", "gridwall shelf black 12pc"),
        ("Huhole-PS-Black-12", "黑色孔板支架12PC", "pegboard shelf black 12pc")):
    p = db.query(Product).filter(Product.sku == sku).first()
    if p is not None:
        p.name_customs_cn = name_cn
        p.name_customs_en = name_en
db.commit()
print("  HUHOLE 盈和产品:", n)

n = import_plan_products(
    CASE + r"\BFPeaky-US-6.5\6.5-US-补仓计划.xlsx",
    CASE + r"\BFPeaky-US-6.5\BF报关价格试算2026.xlsx", "6.5BF-US", BF_DECLARE)
print("  BFPeaky 6.5 产品:", n)

# ================================================================ 3. 模板
print("== 模板 ==")
FWD_K = db.query(Forwarder).filter(Forwarder.name.like("%跨运通%")).first()
FWD_Y = db.query(Forwarder).filter(Forwarder.name.like("%盈和%")).first()
print("  货代: 跨运通=%s 盈和=%s" % (FWD_K.id, FWD_Y.id))


def store_template(src, stored_name):
    dst = os.path.join(TEMPLATE_STORE, stored_name)
    if not os.path.exists(dst):
        shutil.copy2(src, dst)
    return stored_name


def make_yinghe_blank():
    """盈和托书空白模板：从成品复制并清掉数据格（案例里没有空白盈和模板）。"""
    src = (CASE + r"\HUHOLE-US-6.5\HUHOLE-US-6.5\HUHOLE-US-6.5盈和"
           r"\HUHOLE-US-6.5托书1\HUHOLE-普船-ABE8-06.05-托书.xlsx")
    dst = os.path.join(TEMPLATE_STORE, "accept2_盈和托书-空白模板.xlsx")
    if os.path.exists(dst):
        return "accept2_盈和托书-空白模板.xlsx"
    wb = load_workbook(src)
    ws = wb["Sheet1"]
    for addr in ("E4", "L4", "M4", "N4", "E6"):
        ws[addr] = None
    for row in ws.iter_rows(min_row=23, max_row=ws.max_row):
        for c in row:
            c.value = None
    wb.save(dst)
    wb.close()
    return "accept2_盈和托书-空白模板.xlsx"


def make_cangku_blank():
    """投保仓库货值模板：从成品复制清掉数据行。"""
    src = (CASE + r"\HUHOLE-US-6.12\HUHOLE-US-6.12-投保"
           r"\HUHOLE-US-投保仓库货值-0612.xlsx")
    dst = os.path.join(TEMPLATE_STORE, "accept2_投保仓库货值模板.xlsx")
    if os.path.exists(dst):
        return "accept2_投保仓库货值模板.xlsx"
    wb = load_workbook(src)
    ws = wb["Sheet1"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.value = None
    wb.save(dst)
    wb.close()
    return "accept2_投保仓库货值模板.xlsx"


def upsert_template(name, **kw):
    t = db.query(Template).filter(Template.name == name).first()
    if t is None:
        t = Template(name=name)
        db.add(t)
    for k, v in kw.items():
        setattr(t, k, v)
    db.flush()
    return t


# ---- 共用映射 ----
KYT_TUOSHU_MAP = {"sheets": [{
    "name": "运单信息",
    "cells": {
        "F3": "shipment.carton_num",
        "M6": "shipment.postal_code",
        "H7": "{shipment.fc_code}  -  {shipment.address_line1}  -  "
              "{shipment.city}  -  {shipment.state}",
        "H8": "shipment.city",
        "M8": "shipment.state",
        "M9": "shipment.fc_code",
    },
    "tables": [{
        "source": "items", "start_row": 12,
        "columns": {
            "A": "shipment.amazon_shipment_id", "B": "shipment.reference_id",
            "C": "item.carton_lwh_in", "D": "item.box_count",
            "E": "item.box_weight_kg", "F": "item.hs_code",
            "G": "item.name_customs_cn", "H": "item.name_customs_en",
            "I": "item.qty_per_box", "J": "num:2",
            "K": "item.brand", "L": "item.msku", "M": "item.material",
            "N": "item.usage", "O": "item.image_url",
        },
    }],
}]}

YINGHE_TUOSHU_MAP = {"sheets": [{
    "name": "Sheet1",
    "anchors": [
        {"find": "FBA仓库代码*", "in_column": "D", "offset_col": "E",
         "value": "shipment.fc_code"},
        {"find": "客户单号*", "in_column": "D", "offset_col": "E",
         "value": "shipment.amazon_shipment_id"},
        {"find": "客户单号*", "in_column": "D", "offset_col": "L",
         "value": "shipment.amazon_shipment_id"},
        {"find": "客户单号*", "in_column": "D", "offset_col": "M",
         "value": "shipment.reference_id"},
        {"find": "客户单号*", "in_column": "D", "offset_col": "N",
         "value": "shipment.carton_num"},
    ],
    "tables": [{
        "source": "items", "start_row": 23,
        "columns": {
            "B": "item.box_count", "C": "item.name_customs_cn",
            "D": "item.name_customs_en", "E": "item.purchase_cost",
            "F": "item.qty", "G": "item.material", "H": "item.material_en",
            "J": "item.msku", "R": "item.hs_code",
            "U": "text:否", "V": "text:否", "W": "text:否", "X": "num:0",
        },
    }],
}]}

BAOGUAN_MAP = {"sheets": [
    {"name": "报关单",
     "cells": {"A10": "calc.doc_no", "K4": "meta.base_friday_dt"},
     "tables": [{
         "source": "items", "start_row": 20, "row_step": 3,
         "insert_rows": True,
         "columns": {"A": "item.seq", "B": "item.hs_code",
                     "D": "item.name_customs_cn",
                     "D+1": "item.declare_elements",
                     "G": "item.qty", "I": "item.customs_unit_price"},
     }]},
    {"name": "合同",
     "cells": {"F8": "meta.base_friday_dt"},
     "tables": [{
         "source": "items", "start_row": 18, "insert_rows": True,
         "columns": {"A": "item.name_customs_cn", "C": "item.qty",
                     "E": "item.customs_unit_price"},
     }],
     "anchors": [{"find": "总      值", "in_column": "E", "offset_col": "G",
                  "value": "shipment.total_value"}]},
    {"name": "发票",
     "cells": {"G3": "meta.base_friday_dt"},
     "tables": [{
         "source": "items", "start_row": 6, "insert_rows": True,
         "columns": {"C": "item.msku", "D": "item.qty",
                     "F": "item.customs_unit_price"},
     }],
     "anchors": [{"find": "TOTAL:", "in_column": "F", "offset_col": "H",
                  "value": "shipment.total_value"}]},
    {"name": "箱单",
     "tables": [{
         "source": "items", "start_row": 8, "insert_rows": True,
         "columns": {"A": "item.seq", "B": "item.msku",
                     "C": "item.box_count", "D": "item.qty",
                     "F": "item.total_gross_weight"},
     }],
     "anchors": [
         {"find": "合计", "in_column": "A", "offset_col": "C",
          "value": "shipment.carton_num"},
         {"find": "合计", "in_column": "A", "offset_col": "D",
          "value": "shipment.total_qty"},
         {"find": "合计", "in_column": "A", "offset_col": "F",
          "value": "shipment.total_gross_weight"},
         {"find": "合计", "in_column": "A", "offset_col": "G",
          "value": "shipment.total_net_weight"},
     ]},
    {"name": "用途功能",
     "tables": [{
         "source": "items", "start_row": 2, "insert_rows": True,
         "columns": {"A": "item.box_count", "B": "item.qty",
                     "C": "item.msku", "D": "item.name_customs_cn",
                     "E": "item.brand", "F": "item.brand"},
     }]},
]}

TOUBAO_MAP = {"sheets": [{
    "name": "批量导入投保数据",
    "cells": {
        "D3": "shipment.forwarder_order_no",
        "F3": "shipment.forwarder_order_no",
        "G3": "shipment.forwarder_order_no",
        "S3": "shipment.fc_code",
        "U3": "meta.base_friday_slash",
        "W3": "车库挂钩+{shipment.total_qty_float}pcs",
        "Y3": "shipment.carton_num",
        "Z3": "shipment.total_gross_weight",
        "AH3": "shipment.total_contract_value",
    },
}]}


def caigou_map(buyer_ns):
    """采购合同映射；buyer_ns='company'(HUHOLE 买方=舟峰) 或 'trade'(BFPeaky 买方=星盟)。"""
    buyer = {
        "B7": f"{buyer_ns}.name_cn",
        "B9": f"{buyer_ns}.address_cn",
        "B11": f"{buyer_ns}.phone",
    }
    return {"sheets": [{
        "name": "Sheet1",
        "cells": dict({
            "B2": "factory.name", "B3": "factory.address",
            "B5": "factory.phone",
            "F5": "calc.purchase_no", "F7": "batch.contract_date_dt",
        }, **buyer),
        "tables": [{
            "source": "items_agg_by_product", "start_row": 17,
            "insert_rows": True,
            "columns": {"A": "item.name_customs_cn", "C": "item.qty",
                        "E": "calc.price_contract"},
        }],
        "anchors": [{"find": "(9)付款方式", "in_column": "A", "offset_col": "A",
                     "row_offset": -2,
                     "value": "(7) 交货时间：{meta.base_friday}"}],
    }]}


CANGKU_MAP = {"sheets": [{
    "name": "Sheet1",
    "cells": {"A2": "batch.brand"},
    "tables": [{
        "source": "shipments", "start_row": 2,
        "columns": {"B": "shipment.fc_code",
                    "C": "shipment.total_contract_value",
                    "D": "shipment.amazon_shipment_id"},
    }],
}]}

HU65 = CASE + r"\HUHOLE-US-6.5"
BF65 = CASE + r"\BFPeaky-US-6.5"

t7 = upsert_template(
    "新版-跨运通托书(HUBFSE)", owner_type="forwarder", forwarder_id=FWD_K.id,
    doc_type="托书",
    stored_file=store_template(HU65 + r"\拖书模板.xlsx", "accept2_跨运通拖书模板.xlsx"),
    original_name="拖书模板.xlsx", granularity="shipment",
    mapping=json.dumps(KYT_TUOSHU_MAP, ensure_ascii=False),
    filename_rule="{brand}-普船-{fc}-{meta.base_friday_mm_dd}-托书",
    requires_forwarder_no=False, active=True)

t8 = upsert_template(
    "盈和-托书", owner_type="forwarder", forwarder_id=FWD_Y.id, doc_type="托书",
    stored_file=make_yinghe_blank(),
    original_name="盈和托书模板(从成品清空数据制成).xlsx", granularity="shipment",
    mapping=json.dumps(YINGHE_TUOSHU_MAP, ensure_ascii=False),
    filename_rule="{brand}-普船-{fc}-{meta.base_friday_mm_dot_dd}-托书",
    requires_forwarder_no=False, active=True)

t9 = upsert_template(
    "HUHOLE-报关资料", owner_type="forwarder", forwarder_id=FWD_K.id,
    doc_type="报关资料",
    stored_file=store_template(HU65 + r"\HU报关模板(新）.xlsx", "accept2_HU报关模板.xlsx"),
    original_name="HU报关模板(新）.xlsx", granularity="shipment",
    mapping=json.dumps(BAOGUAN_MAP, ensure_ascii=False),
    filename_rule="{brand}-{country}-{fc}-报关资料-{shipdate}",
    requires_forwarder_no=False, active=True)

t10 = upsert_template(
    "HUHOLE-投保", owner_type="forwarder", forwarder_id=FWD_K.id, doc_type="投保单",
    stored_file=store_template(HU65 + r"\投保模板.xlsx", "accept2_HU投保模板.xlsx"),
    original_name="投保模板.xlsx", granularity="shipment",
    mapping=json.dumps(TOUBAO_MAP, ensure_ascii=False),
    filename_rule="{brand}-投保-{meta.base_friday_mm_dd}-{fc}",
    requires_forwarder_no=True, active=True)

t11 = upsert_template(
    "HUHOLE-采购合同", owner_type="internal", doc_type="采购合同",
    stored_file=store_template(HU65 + r"\工厂-舟峰-采购合同模板(1).xlsx",
                               "accept2_HU采购合同模板.xlsx"),
    original_name="工厂-舟峰-采购合同模板(1).xlsx", granularity="batch",
    mapping=json.dumps(caigou_map("company"), ensure_ascii=False),
    filename_rule="{brand}-采购合同-{shipdate}",
    requires_forwarder_no=False, active=True)

t12 = upsert_template(
    "BFPeaky-报关资料", owner_type="forwarder", forwarder_id=FWD_K.id,
    doc_type="报关资料",
    stored_file=store_template(BF65 + r"\BF报关模板(新）.xlsx", "accept2_BF报关模板.xlsx"),
    original_name="BF报关模板(新）.xlsx", granularity="shipment",
    mapping=json.dumps(BAOGUAN_MAP, ensure_ascii=False),
    filename_rule="{brand}-{country}-{fc}-报关资料-{shipdate}",
    requires_forwarder_no=False, active=True)

t13 = upsert_template(
    "BFPeaky-投保", owner_type="forwarder", forwarder_id=FWD_K.id, doc_type="投保单",
    stored_file=store_template(BF65 + r"\投保模板.xlsx", "accept2_BF投保模板.xlsx"),
    original_name="投保模板.xlsx", granularity="shipment",
    mapping=json.dumps(TOUBAO_MAP, ensure_ascii=False),
    filename_rule="{brand}-投保-{meta.base_friday_mm_dd}-{fc}",
    requires_forwarder_no=True, active=True)

t14 = upsert_template(
    "BFPeaky-采购合同", owner_type="internal", doc_type="采购合同",
    stored_file=store_template(BF65 + r"\BF采购合同模板.xlsx",
                               "accept2_BF采购合同模板.xlsx"),
    original_name="BF采购合同模板.xlsx", granularity="batch",
    mapping=json.dumps(caigou_map("trade"), ensure_ascii=False),
    filename_rule="{brand}-采购合同-{shipdate}",
    requires_forwarder_no=False, active=True)

t15 = upsert_template(
    "HUBFSE-投保仓库货值", owner_type="internal", doc_type="投保单",
    stored_file=make_cangku_blank(),
    original_name="投保仓库货值模板(从成品清空数据制成).xlsx",
    granularity="row_per_shipment",
    mapping=json.dumps(CANGKU_MAP, ensure_ascii=False),
    filename_rule="{brand}-{country}-投保仓库货值-{shipdate}",
    requires_forwarder_no=False, active=True)

db.commit()
print("  templates:", [(t.id, t.name) for t in
                       (t7, t8, t9, t10, t11, t12, t13, t14, t15)])
db.close()
print("done")

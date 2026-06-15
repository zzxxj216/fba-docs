"""集成端到端验证：补 Serenorch 品牌链 → 绑定批次 → 校验 → 上传模板 → 生成。可重复运行。"""
import io
import json

import httpx
from openpyxl import Workbook, load_workbook

BASE = "http://127.0.0.1:8000/api"
c = httpx.Client(timeout=120)


def api(method, path, **kw):
    r = c.request(method, BASE + path, **kw)
    if r.status_code >= 400:
        raise RuntimeError(f"{method} {path} -> {r.status_code}: {r.text[:300]}")
    return r.json() if "json" in r.headers.get("content-type", "") else r.content


def find_or_create(resource, match_key, match_val, payload):
    for row in api("GET", f"/{resource}"):
        if row.get(match_key) == match_val:
            return row
    return api("POST", f"/{resource}", json=payload)


# 1. Serenorch 业务链（工厂信息库里没有，手动补——正是系统设计支持的场景）
factory = find_or_create("factories", "name", "安徽安庆市嘉欣医疗用品科技股份有限公司", {
    "name": "安徽安庆市嘉欣医疗用品科技股份有限公司", "abbr": "JX",
    "address": "安徽省安庆市宜秀区胜利路76号", "phone": "17755682899",
})
company = find_or_create("companies", "name_cn", "玫玑研日用品（上海）有限公司", {
    "type": "shop", "name_cn": "玫玑研日用品（上海）有限公司",
    "name_en": "Mei Ji Yan Ri Yong Pin (Shanghai) Co., Ltd.",
    "address_cn": "上海市普陀区绥德路2弄20号5层A503室",
    "address_en": "A503 Shi 5 Ceng 20 Hao 2 Nong Suide Lu Putuo Qu Shanghai Shi China",
    "uscc": "91310107MA1G0D9E08", "phone": "19116409320",
    "insurance_factor": 1.0, "export_via_trade": False,
    "default_port": "宁波", "default_origin_place": "杭州",
})
brand = find_or_create("brands", "name", "Serenorch", {
    "name": "Serenorch", "abbr2": "SE",
    "factory_id": factory["id"], "company_id": company["id"],
    "doc_no_rule_shipment": "{sa}{brand2}-{fc}-{date}",
    "doc_no_rule_purchase": "{sa}{brand2}-{country}-{date}",
})
print(f"1. 业务链就绪: 工厂#{factory['id']} 主体#{company['id']} 品牌#{brand['id']}")

# 2. 绑定批次
api("PUT", "/batches/1", json={
    "brand_id": brand["id"], "company_id": company["id"], "factory_id": factory["id"],
})
report = api("GET", "/batches/1/validate")
print(f"2. 绑定后校验: passed={report['passed']}")
for e in report["errors"]:
    print(f"   [{e['level']}] {e['msg']}")
assert report["passed"], "校验未通过，无法继续生成"

# 3. 造一个迷你托书测试模板并上传（锚点+坐标+明细区三种定位都覆盖）
wb = Workbook()
ws = wb.active
ws.title = "托书"
ws["A2"], ws["A3"], ws["A4"] = "收件仓:", "FBA货件号:", "Reference:"
ws["C2"], ws["C3"] = "总箱数:", "总毛重kg:"
ws["A6"] = "客户单号"  # 锚点：往 C 列填发货单号
ws["A8"], ws["B8"], ws["C8"], ws["D8"], ws["E8"] = "SKU", "英文报关名", "数量", "箱数", "金额"
buf = io.BytesIO()
wb.save(buf)
up = api("POST", "/templates/upload-file",
         files={"file": ("测试托书模板.xlsx", buf.getvalue(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
print(f"3. 模板上传: {up['stored_file']} sheets={up['sheetnames']}")

mapping = {"sheets": [{
    "name": "托书",
    "cells": {
        "B2": "shipment.fc_code",
        "B3": "shipment.amazon_shipment_id",
        "B4": "shipment.reference_id",
        "D2": "shipment.carton_num",
        "D3": "shipment.total_gross_weight",
        "F2": "calc.doc_no",
    },
    "anchors": [
        {"find": "客户单号", "in_column": "A", "offset_col": "C", "value": "shipment.ship_sn"},
    ],
    "tables": [{
        "source": "items", "start_row": 9, "insert_rows": True,
        "columns": {"A": "item.msku", "B": "item.name_customs_en",
                    "C": "item.qty", "D": "item.box_count", "E": "item.amount"},
    }],
}]}
tpl = None
for row in api("GET", "/templates"):
    if row.get("name") == "集成测试-托书":
        tpl = api("PUT", f"/templates/{row['id']}", json={
            "stored_file": up["stored_file"], "original_name": up["original_name"],
            "mapping": json.dumps(mapping)})
        break
if tpl is None:
    tpl = api("POST", "/templates", json={
        "name": "集成测试-托书", "owner_type": "forwarder", "doc_type": "托书",
        "stored_file": up["stored_file"], "original_name": up["original_name"],
        "granularity": "shipment", "mapping": json.dumps(mapping),
        "filename_rule": "测试托书-{fc}-{MMDD}",
    })
print(f"4. 模板配置: id={tpl['id']}")

# 4. 生成
result = api("POST", "/batches/1/generate", json={"template_ids": [tpl["id"]]})
print(f"5. 生成结果: {json.dumps(result, ensure_ascii=False)[:400]}")
assert not result.get("blocked"), "被校验拦截"
gen = result.get("generated") or []
assert len(gen) == 5, f"应生成5份(每货件一份)，实际 {len(gen)}"

# 5. 下载第一份验证内容
content = api("GET", f"/docs/{gen[0]['doc_id']}/download")
out = load_workbook(io.BytesIO(content))
w = out["托书"]
print("6. 文件内容抽查:")
print(f"   B2 仓={w['B2'].value}  B3 FBA={w['B3'].value}  B4 Ref={w['B4'].value}")
print(f"   D2 箱数={w['D2'].value}  D3 毛重={w['D3'].value}  F2 编号={w['F2'].value}")
print(f"   C6 锚点发货单号={w['C6'].value}")
print(f"   明细第9行: {w['A9'].value} | {w['B9'].value} | {w['C9'].value} | {w['D9'].value} | {w['E9'].value}")
print(f"   明细第10行: {w['A10'].value} | {w['B10'].value} | {w['C10'].value} | {w['D10'].value}")
assert w["B3"].value and str(w["B3"].value).startswith("FBA")
assert w["C6"].value and str(w["C6"].value).startswith("FH")
assert w["A9"].value and w["A10"].value  # 两个 SKU 都在

# 6. zip
z = api("GET", "/batches/1/zip")
assert z[:2] == b"PK"
print(f"7. 打包下载 OK ({len(z)} bytes)")
print("\n端到端集成验证全部通过 ✅")

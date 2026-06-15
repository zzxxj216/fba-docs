# -*- coding: utf-8 -*-
"""Agent B 生成引擎冒烟测试：rule_engine / excel_engine / field_registry / ai_mapping。

直接运行：python _smoke_b.py（需要 MySQL 就绪；不向库里写业务数据，只种规则默认值）。
"""

import os
import sys
import tempfile
from datetime import date

sys.stdout.reconfigure(encoding="utf-8")

from app.database import Base, engine, SessionLocal  # noqa: E402

Base.metadata.create_all(engine)

PASS = 0


def ok(cond, msg):
    global PASS
    assert cond, f"FAIL: {msg}"
    PASS += 1
    print(f"  ok - {msg}")


# ================================================================ ① rule_engine
print("① rule_engine")
from app import rule_engine as re_  # noqa: E402

ok(re_.next_friday(date(2026, 6, 12)) == date(2026, 6, 12), "next_friday: 周五(06-12)→当天")
ok(re_.next_friday(date(2026, 6, 8)) == date(2026, 6, 12), "next_friday: 周一(06-08)→本周五06-12")
ok(re_.next_friday(date(2026, 6, 11)) == date(2026, 6, 12), "next_friday: 周四(06-11)→本周五")
ok(re_.next_friday(date(2026, 6, 13)) == date(2026, 6, 19), "next_friday: 周六(06-13)→下周五06-19")
ok(re_.next_friday(date(2026, 6, 14)) == date(2026, 6, 19), "next_friday: 周日(06-14)→下周五06-19")
ok(re_.next_friday("2026-06-08") == date(2026, 6, 12), "next_friday: 接受字符串日期")

ok(re_.prev_month_same_day(date(2026, 6, 5)) == date(2026, 5, 5), "prev_month_same_day: 06-05→05-05")
ok(re_.prev_month_same_day(date(2026, 3, 31)) == date(2026, 2, 28),
   "prev_month_same_day: 03-31→02-28（2026非闰年，上月无31取月末）")
ok(re_.prev_month_same_day(date(2024, 3, 31)) == date(2024, 2, 29),
   "prev_month_same_day: 2024-03-31→02-29（闰年）")
ok(re_.prev_month_same_day(date(2026, 5, 31)) == date(2026, 4, 30), "prev_month_same_day: 05-31→04-30")
ok(re_.prev_month_same_day(date(2026, 1, 15)) == date(2025, 12, 15), "prev_month_same_day: 跨年 01-15→上年12-15")

ok(re_.net_weight(100, 10) == 95.0, "net_weight: 100kg/10箱→95.0（默认扣0.5）")
ok(re_.price_vat(10) == 11.3, "price_vat: 10→11.3")
ok(re_.price_vat_markup(10) == 12.43, "price_vat_markup: 10→12.43")


class _O:  # 简易对象
    def __init__(self, **kw):
        self.__dict__.update(kw)


_brand = _O(abbr2="SE", name="Serenorch",
            doc_no_rule_shipment="{sa}{brand2}-{fc}-{date}",
            doc_no_rule_purchase="{sa}{brand2}-{country}-{date}")
_co_sa = _O(export_via_trade=True, insurance_factor=1.13, id=None)
_co_no = _O(export_via_trade=False, insurance_factor=1.0, id=None)
ok(re_.doc_no(_brand, _co_sa, "ONT8", "2026-05-19", "shipment") == "SA-SE-ONT8-2026-05-19",
   "doc_no: {date}=全日期 → SA-SE-ONT8-2026-05-19")
ok(re_.doc_no(_brand, _co_no, "ONT8", "2026-05-19", "purchase", country="US") == "SE-US-2026-05-19",
   "doc_no: 一级采购合同 SE-US-2026-05-19")
ok(re_.doc_no(_brand, _co_no, "ABE8", "2026-05-06", "shipment") == "SE-ABE8-2026-05-06",
   "doc_no: 复刻历史 9810 订单编号 SE-ABE8-2026-05-06")
_brand_mmdd = _O(abbr2="SE", name="Serenorch",
                 doc_no_rule_shipment="{sa}{brand2}-{fc}-{date_mmdd}",
                 doc_no_rule_purchase="{sa}{brand2}-{country}-{date_mmdd}")
ok(re_.doc_no(_brand_mmdd, _co_no, "ONT8", "2026-05-19", "shipment") == "SE-ONT8-0519",
   "doc_no: {date_mmdd} 保留月日形态 SE-ONT8-0519")
ok(re_.insurance_price(2.0, _co_sa) == 2.26, "insurance_price: 2.0×1.13=2.26")

# ================================================================ ② excel_engine
print("② excel_engine")
from openpyxl import Workbook, load_workbook  # noqa: E402
from app import excel_engine  # noqa: E402

tmp_dir = tempfile.mkdtemp(prefix="smoke_b_")
tpl_path = os.path.join(tmp_dir, "tpl.xlsx")
out_path = os.path.join(tmp_dir, "out.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "模板"
ws["A1"] = "托书"
ws["A2"] = "目的仓："
ws["A5"] = "客户单号"           # 锚点行
ws["A17"] = "序号"; ws["B17"] = "SKU"; ws["C17"] = "数量"  # 明细表头
ws["A18"] = None                 # 明细起始行
ws["A18"].font = ws["A18"].font.copy(bold=True)  # 给起始行一点样式供复制
wb.create_sheet("说明页")
wb.save(tpl_path)

fake_ctx = {
    "shipment": {"fc_code": "ONT8", "ship_sn": "FH20260612", "carton_num": 13},
    "items": [
        {"item": {"seq": 1, "msku": "SKU1", "qty": 10}, "calc": {"price_vat": 11.3}},
        {"item": {"seq": 2, "msku": "SKU2", "qty": 20}, "calc": {"price_vat": 22.6}},
        {"item": {"seq": 3, "msku": "SKU3", "qty": 30}, "calc": {"price_vat": 33.9}},
    ],
    "meta": {"base_friday_mmdd": "0612"},
}
mapping = {"sheets": [{
    "name": "模板",
    "cells": {"B2": "shipment.fc_code", "D2": "text:海运", "F2": "{shipment.fc_code}-{meta.base_friday_mmdd}"},
    "anchors": [{"find": "客户单号", "in_column": "A", "offset_col": "L",
                 "row_offset": 0, "value": "shipment.ship_sn"}],
    "tables": [{"source": "items", "start_row": 18, "insert_rows": True,
                "columns": {"A": "item.seq", "B": "item.msku",
                            "C": "item.qty", "D": "calc.price_vat",
                            "E": "{item.msku}×{item.qty}"}}],
}]}
excel_engine.fill(tpl_path, mapping, fake_ctx, out_path)

rb = load_workbook(out_path)
rs = rb["模板"]
ok(rs["B2"].value == "ONT8", "cells: B2=ONT8（固定坐标）")
ok(rs["D2"].value == "海运", "cells: text: 前缀字面量")
ok(rs["F2"].value == "ONT8-0612", "cells: 格式串拼接")
ok(rs["L5"].value == "FH20260612", "anchors: 客户单号行 L 列=发货单号")
ok([rs[f"A{r}"].value for r in (18, 19, 20)] == [1, 2, 3], "tables: 三行序号 1/2/3（插行）")
ok(rs["B19"].value == "SKU2" and rs["C20"].value == 30, "tables: 行内 item.* 路由到当前行")
ok(isinstance(rs["C18"].value, int) and isinstance(rs["D18"].value, float),
   "tables: 数字保持数值类型")
ok(rs["E20"].value == "SKU3×30", "tables: 行内格式串")
ok(rs["A19"].font.bold and rs["A20"].font.bold, "tables: 插入行复制了起始行样式")

# 错误信息可读性
try:
    excel_engine.fill(tpl_path, {"sheets": [{"name": "不存在的表", "cells": {}}]},
                      fake_ctx, out_path)
    ok(False, "缺工作表应抛 ValueError")
except ValueError as e:
    ok("不存在的表" in str(e), f"缺工作表报中文错误：{e}")
try:
    excel_engine.fill(tpl_path, {"sheets": [{"name": "模板",
                      "cells": {"B3": "shipment.no_such_field"}}]}, fake_ctx, out_path)
    ok(False, "未登记字段应抛 ValueError")
except ValueError as e:
    ok("B3" in str(e) and "no_such_field" in str(e), f"坏路径报哪个格子哪个路径：{e}")

# ---- row_step：一条记录占多行（报关单 D20品名/D21申报要素/G21单位/I22币制）
tpl3_path = os.path.join(tmp_dir, "tpl_customs.xlsx")
out3_path = os.path.join(tmp_dir, "out_customs.xlsx")
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "报关单"
ws3["A19"] = "项号"; ws3["D19"] = "商品名称及规格型号"
ws3["A26"] = "录入员"      # 明细区下方的固定布局，不许被推乱
wb3.save(tpl3_path)

fake3 = {"items": [
    {"item": {"seq": 1, "name": "一次性面巾纸", "declare": "用途：盥洗用|材质：无纺布",
              "unit": "盒", "currency": "USD"}},
    {"item": {"seq": 2, "name": "化妆棉", "declare": "用途：盥洗用|材质：无纺布2",
              "unit": "盒", "currency": "USD"}},
]}
mapping3 = {"sheets": [{
    "name": "报关单",
    "tables": [{"source": "items", "start_row": 20, "row_step": 3,
                # insert_rows 缺省=False：模板自带空行，插行会破坏 A26 固定落款
                "columns": {"A": "item.seq", "D": "item.name",
                            "D+1": "item.declare", "G+1": "item.unit",
                            "I+2": "item.currency"}}],
}]}
excel_engine.fill(tpl3_path, mapping3, fake3, out3_path)
rb3 = load_workbook(out3_path)
rs3 = rb3["报关单"]
ok(rs3["D20"].value == "一次性面巾纸" and rs3["A20"].value == 1,
   "row_step: 第1条 D20=品名")
ok(rs3["D21"].value == "用途：盥洗用|材质：无纺布", "row_step: 第1条 D+1→D21=申报要素")
ok(rs3["G21"].value == "盒" and rs3["I22"].value == "USD",
   "row_step: 第1条 G+1→G21单位、I+2→I22币制")
ok(rs3["D23"].value == "化妆棉" and rs3["D24"].value == "用途：盥洗用|材质：无纺布2",
   "row_step: 第2条整体下移3行（D23/D24）")
ok(rs3["G24"].value == "盒" and rs3["I25"].value == "USD", "row_step: 第2条 G24/I25")
ok(rs3["A26"].value == "录入员", "insert_rows 默认 False：明细区下方固定布局未被推乱")
try:
    excel_engine.fill(tpl3_path, {"sheets": [{"name": "报关单",
                      "tables": [{"source": "items", "start_row": 20,
                                  "columns": {"D+x": "item.name"}}]}]},
                      fake3, out3_path)
    ok(False, "非法列名应抛 ValueError")
except ValueError as e:
    ok("D+x" in str(e), f"非法列名报中文错误：{e}")

# ================================================================ ③ field_registry.build_context
print("③ field_registry.build_context（内存 ORM 对象 + 真库读规则）")
from app.models import (Batch, Brand, Company, Factory, Product,  # noqa: E402
                        Shipment, ShipmentBox, ShipmentItem)
from app.field_registry import build_context, resolve  # noqa: E402
from app.seed import seed_rules  # noqa: E402

db = SessionLocal()
seed_rules(db)  # 保证规则默认值在位（幂等）

product = Product(sku="SR-FT-01", name_customs_cn="足部按摩垫", name_customs_en="Foot Massage Mat",
                  name_contract="按摩垫", name_invoice="Massage Mat", name_usage="按摩放松",
                  hs_code="9019101000", material="ABS+布", usage="按摩", brand_name="Serenorch",
                  model="FT-01", box_weight_kg=5.0, image_url="http://img/x.jpg")
product2 = Product(sku="SR-CP-01", name_customs_cn="化妆棉", name_customs_en="Cotton Pad",
                   name_contract="化妆棉", name_invoice="Cotton Pad", name_usage="化妆棉1pc",
                   hs_code="6302930010", material="无纺布", usage="盥洗用", brand_name="Serenorch",
                   model="", box_weight_kg=8.25)   # 无型号 → declare_full 不带 |规格：
factory = Factory(name="临公工厂", abbr="LG", address="杭州萧山", contact="王厂", phone="139")
company = Company(type="shop", name_cn="杭州某某科技", name_en="Hangzhou XX Tech",
                  abbr="XX", insurance_factor=1.13, export_via_trade=False,
                  default_port="宁波", default_origin_place="萧山")
brand = Brand(name="Serenorch", abbr2="SE")
item = ShipmentItem(msku="SR-FT-01", fnsku="X001ABC", asin="B0TEST", qty=100,
                    box_count=10, qty_per_box=10, carton_l=50.0, carton_w=40.0,
                    carton_h=30.0, customs_unit_price=2.0, purchase_cost=10.0)
item.product = product
box1 = ShipmentBox(box_id="FBA17U000001", msku="SR-FT-01", qty=10,
                   length_in=20.0, width_in=16.0, height_in=12.0, weight_lb=22.0)
box2 = ShipmentBox(box_id="FBA17U000002", msku="SR-FT-01", qty=10,
                   length_in=20.0, width_in=16.0, height_in=12.0, weight_lb=22.0)
sp = Shipment(amazon_shipment_id="FBA17XYZ", ship_sn="FH2026061201", reference_id="REF1",
              fc_code="ONT8", address_line1="2 Saint Road", city="San Bernardino",
              state="CA", postal_code="92408", carton_num=10,
              forwarder_order_no="XX177")
sp.items = [item]
sp.boxes = [box1, box2]
# 第二货件：同 SKU 再 50（验证 items_agg 跨货件聚合）+ 另一 SKU（验证 plan_items 行数）
item2 = ShipmentItem(msku="SR-FT-01", qty=50, box_count=5, qty_per_box=10,
                     customs_unit_price=2.0, purchase_cost=10.0)
item2.product = product
item3 = ShipmentItem(msku="SR-CP-01", qty=30, box_count=3, qty_per_box=10,
                     carton_l=54.0, carton_w=42.0, carton_h=42.0,
                     customs_unit_price=1.5, purchase_cost=7.0)
item3.product = product2
sp2 = Shipment(amazon_shipment_id="FBA17ABC", ship_sn="FH2026061202", reference_id="REF2",
               fc_code="ABE8", city="Allentown", state="PA", postal_code="18031",
               carton_num=8)
sp2.items = [item2, item3]
sp2.boxes = []
batch = Batch(name="Serenorch-US-0612", country="US", marketplace="US",
              base_date="2026-06-12", contract_date="2026-05-12", shop_name="Serenorch-US")
batch.brand = brand
batch.company = company
batch.factory = factory
batch.shipments = [sp, sp2]

ctx = build_context(db, batch, sp)
ok(ctx["shipment"]["total_gross_weight"] == 50.0, "总毛重=10箱×5kg=50.0")
ok(ctx["shipment"]["total_net_weight"] == 45.0, "总净重=50−10×0.5=45.0（走规则表）")
ok(ctx["shipment"]["total_qty"] == 100 and ctx["shipment"]["total_value"] == 200.0,
   "总数量100 / 总金额200.0")
row = ctx["items"][0]
ok(row["item"]["amount"] == 200.0 and row["item"]["name_customs_cn"] == "足部按摩垫",
   "item 行：amount=qty×单价，品名取自 Product")
ok(row["calc"]["price_vat"] == 11.3 and row["calc"]["price_vat_markup"] == 12.43,
   "行级 calc：含税 11.3 / 二级加价 12.43")
ok(row["calc"]["insurance_price"] == 2.26 and row["calc"]["insurance_amount"] == 226.0,
   "行级 calc：保价 2.26 / 226.0（×公司保价倍率1.13）")
ok(row["item"]["carton_l_in"] == 19.69 and row["item"]["carton_w_in"] == 15.75,
   "item 箱规英寸：50cm→19.69in / 40cm→15.75in（÷2.54 圆整2位）")
ok(row["item"]["box_weight_kg"] == 5.0 and row["item"]["total_gross_weight"] == 50.0,
   "item 单箱重量 5.0kg / 总毛重=5.0×10箱=50.0")
ok(row["item"]["declare_full"] == "用途：按摩|材质：ABS+布|品牌：Serenorch|规格：FT-01",
   "declare_full：有型号带 |规格：（对照 RPA 报关单 D21 串）")
b = ctx["boxes"][0]["box"]
ok(b["box_no_text"] == "1/2" and ctx["boxes"][1]["box"]["box_no_text"] == "2/2",
   "box_no_text 连续编号 1/2、2/2")
ok(b["length_cm"] == 50.8 and b["weight_kg"] == 9.98,
   "箱尺寸双单位：20in→50.8cm、22lb→9.98kg")
ok(ctx["calc"]["contract_date_cn"] == "2026年5月12日", "contract_date_cn 中文日期")
ok("SE-ONT8-2026-05-12" in ctx["calc"]["doc_no"],
   f"货件编号 {{date}}=全日期：{ctx['calc']['doc_no']}")
ok("SE-US-2026-05-12" in ctx["calc"]["purchase_no"],
   f"采购编号 {{date}}=全日期：{ctx['calc']['purchase_no']}")
ok(ctx["calc"]["supervision_mode"] == "跨境电商出口海外仓", "报关常量走 RuleConfig")
ok(ctx["meta"]["base_friday"] == "2026-06-12" and ctx["meta"]["base_friday_mmdd"] == "0612",
   "meta.base_friday 多格式")

ok(resolve("shipment.fc_code", ctx) == "ONT8", "resolve 普通路径")
ok(resolve("text:固定文字", ctx) == "固定文字", "resolve text: 前缀")
ok(resolve("{batch.brand}-托书-{shipment.fc_code}", ctx) == "Serenorch-托书-ONT8",
   "resolve 格式串")

# 批次级上下文（row_per_shipment / items_agg / plan_items 用）
bctx = build_context(db, batch)
ok(len(bctx["shipments"]) == 2, "批次级 ctx 含 shipments 行列表（2 货件）")
srow = bctx["shipments"][0]
ok(srow["shipment"]["total_qty"] == 100 and srow["item"]["msku"] == "SR-FT-01",
   "shipments 行=货件汇总+第一明细产品字段")
ok(resolve("shipment.forwarder_order_no", bctx, srow) == "XX177",
   "行内 resolve：shipment.* 路由到当前货件行")

# items_agg：跨货件按 msku 聚合（采购合同=批次总量）
agg = bctx["items_agg"]
ok(len(agg) == 2, "items_agg：3 条明细按 msku 聚成 2 行")
a0 = agg[0]
ok(a0["item"]["msku"] == "SR-FT-01" and a0["item"]["qty"] == 150
   and a0["item"]["box_count"] == 15, "items_agg：qty 100+50=150 / 箱数 10+5=15")
ok(a0["item"]["amount"] == 300.0 and a0["item"]["total_gross_weight"] == 75.0,
   "items_agg：金额/毛重按聚合数量重算（150×2.0=300 / 5kg×15箱=75）")
ok(a0["item"]["carton_l"] == 50.0 and a0["calc"]["price_vat"] == 11.3,
   "items_agg：非求和字段取第一条（箱规/含税单价）")
ok(a0["calc"]["amount_vat"] == 1695.0, "items_agg：含税金额=11.3×150=1695.0")
a1 = agg[1]
ok(a1["item"]["msku"] == "SR-CP-01" and a1["item"]["qty"] == 30 and a1["item"]["seq"] == 2,
   "items_agg：第二 SKU 独立成行，seq 连续编号")
ok(a1["item"]["carton_l_in"] == 21.26, "items_agg/item：54cm÷2.54=21.26in")
ok(a1["item"]["declare_full"] == "用途：盥洗用|材质：无纺布|品牌：Serenorch",
   "declare_full：无型号不带 |规格：（照 RPA 分支）")

# plan_items：每 (货件×SKU) 一行，calc.doc_no 用该货件的（9810）
plan = bctx["plan_items"]
ok(len(plan) == 3, "plan_items：1+2=3 行（货件×SKU）")
ok(plan[0]["calc"]["doc_no"] == "SE-ONT8-2026-05-12"
   and plan[1]["calc"]["doc_no"] == "SE-ABE8-2026-05-12"
   and plan[2]["calc"]["doc_no"] == "SE-ABE8-2026-05-12",
   "plan_items：calc.doc_no 按行所属货件（ONT8/ABE8）")
ok(plan[2]["item"]["msku"] == "SR-CP-01" and plan[2]["item"]["qty"] == 30
   and plan[2]["calc"]["price_vat"] == 7.91,
   "plan_items：行级 item/calc 是该 SKU 自己的（7×1.13=7.91）")
ok(plan[1]["shipment"]["fc_code"] == "ABE8" and plan[1]["shipment"]["total_qty"] == 80,
   "plan_items：行内 shipment.* 为所属货件汇总")
ok(resolve("calc.doc_no", bctx, plan[1]) == "SE-ABE8-2026-05-12",
   "行内 resolve：plan_items 行 calc.* 路由正确")

# 文件名渲染
from app.services.generate_service import _render_filename  # noqa: E402
_t = _O(filename_rule="{brand}-托书-{fc}-{shipdate}", doc_type="托书")
ok(_render_filename(_t, ctx, batch, sp) == "Serenorch-托书-ONT8-0612.xlsx",
   "filename_rule 短占位符渲染")
_t2 = _O(filename_rule="", doc_type="托书")
ok(_render_filename(_t2, ctx, batch, sp) == "托书-ONT8-0612.xlsx",
   "空规则默认 {doc_type}-{fc}-{MMDD}")

db.close()

# ================================================================ ④ ai_mapping
print("④ ai_mapping（无 API key 时优雅降级）")
from app import ai_mapping  # noqa: E402

TPL = r"C:\Users\zane\Downloads\Serenorch-US-6.5\托书-空白模板.xlsx"
if os.path.exists(TPL):
    text, names = ai_mapping._extract_sheet_text(TPL)
    ok(len(text) > 100 and names, f"模板单元格提取 OK（sheets={names}，{len(text)} 字符）")
    result = ai_mapping.suggest_mapping(TPL)
    if "error" in result:
        ok("ANTHROPIC_API_KEY" in result["error"] or "失败" in result["error"],
           f"无 key/调用失败 → 优雅降级：{result['error']}")
    else:
        ok("mapping" in result and "confidences" in result,
           f"AI 映射成功：{len(result['confidences'])} 条建议")
        print("  notes:", (result.get("notes") or "")[:200])
else:
    ok(True, f"跳过：样例模板不存在 {TPL}")
    r = ai_mapping.suggest_mapping("不存在.xlsx")
    ok("error" in r, "文件不存在 → error 降级")

print(f"\n全部通过：{PASS} 项断言")

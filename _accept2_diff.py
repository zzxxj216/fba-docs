# -*- coding: utf-8 -*-
"""验收2 逐单元格 diff：本系统生成文件 vs RPA 成品（四票）。

用法: python _accept2_diff.py [票名...]   票名 ∈ SE HUK BF HUY，缺省全部
分类: MATCH / 日期类差异 / 数据缺口 / RPA bug / 我方错误
"""
import datetime as _dt
import os
import re
import sys

from openpyxl import load_workbook

CASE = r"C:\Users\zane\Downloads\HUBFSE自动化"
OUT = r"D:\amazon\output"

# ---------------------------------------------------------------- 文件配对
PAIRS = []   # (case, label, ours, theirs)


def pair(case, label, ours, theirs):
    PAIRS.append((case, label, ours, theirs))


# Serenorch-US-6.19（跨运通复验，模板 id2-6）
SE_SRC = CASE + r"\Serenorch-US-6.19"
SE_OUT = OUT + r"\Serenorch-US-0619"
for fc in ["ABE8", "GYR2", "IND9", "MDW2", "SCK4"]:
    pair("SE", f"托书-{fc}",
         rf"{SE_OUT}\{fc}\Serenorch-托书-{fc}-0619.xlsx",
         rf"{SE_SRC}\Serenorch-US-6.19-托书\Serenorch-托书-{fc}-0619.xlsx")
    pair("SE", f"报关-{fc}",
         rf"{SE_OUT}\{fc}\Serenorch-US-{fc}-报关资料-0619.xlsx",
         rf"{SE_SRC}\Serenorch-US-6.19-报关\Serenorch-US-{fc}-报关资料-0619.xlsx")
    pair("SE", f"投保-{fc}",
         rf"{SE_OUT}\{fc}\Serenorch-投保-06-19-{fc}.xlsx",
         rf"{SE_SRC}\Serenorch-US-6.19-投保\Serenorch-投保-06-19-{fc}.xlsx")
pair("SE", "采购合同",
     rf"{SE_OUT}\Serenorch - 采购合同-2026-06-19.xlsx",
     rf"{SE_SRC}\Serenorch-US-6.19-采购合同\Serenorch - 采购合同-2026-06-19.xlsx")
pair("SE", "9810",
     rf"{SE_OUT}\Serenorch - US - 报关9810-20260612.xlsx",
     rf"{SE_SRC}\Serenorch-US-6.19-9810\Serenorch - US - 报关9810-20260619.xlsx")

# HUHOLE-US-6.5 跨运通
HU_SRC = CASE + r"\HUHOLE-US-6.5"
HU_OUT = OUT + r"\HUHOLE-US-0605K"
for fc in ["GYR2", "IND9", "LGB8", "MQJ1", "VGT2"]:
    pair("HUK", f"托书-{fc}",
         rf"{HU_OUT}\{fc}\HUHOLE-普船-{fc}-06-05-托书.xlsx",
         rf"{HU_SRC}\HUHOLE-US-6.5\HUHOLE-US-6.5跨运通\HUHOLE-US-6.5托书"
         rf"\HUHOLE-普船-{fc}-06-02-托书.xlsx")
    pair("HUK", f"报关-{fc}",
         rf"{HU_OUT}\{fc}\HUHOLE-US-{fc}-报关资料-0605.xlsx",
         rf"{HU_SRC}\HUHOLE-US-6.5-报关单\HUHOLE-US-{fc}-报关资料-0605.xlsx")
    pair("HUK", f"投保-{fc}",
         rf"{HU_OUT}\{fc}\HUHOLE-投保-06-05-{fc}.xlsx",
         rf"{HU_SRC}\HUHOLE-US-6.5-投保\HUHOLE-投保-06-05-{fc}.xlsx")
pair("HUK", "采购合同",
     rf"{HU_OUT}\HUHOLE-采购合同-0605.xlsx",
     rf"{HU_SRC}\HUHOLE-US-6.5-报关单\HUHOLE-采购合同-0605.xlsx")

# BFPeaky-US-6.5（抽查；该周无采购合同成品）
BF_SRC = CASE + r"\BFPeaky-US-6.5"
BF_OUT = OUT + r"\BFPeaky-US-0605"
for fc in ["GYR3", "IND9", "RFD2", "RMN3", "SBD1"]:
    pair("BF", f"托书-{fc}",
         rf"{BF_OUT}\{fc}\BFPeaky-普船-{fc}-06-05-托书.xlsx",
         rf"{BF_SRC}\BFPeaky-US-6.5\BFPeaky-US-6.5-托书"
         rf"\BFPeaky-普船-{fc}-06-02-托书.xlsx")
    pair("BF", f"报关-{fc}",
         rf"{BF_OUT}\{fc}\BFPeaky-US-{fc}-报关资料-0605.xlsx",
         rf"{BF_SRC}\BFPeaky-US-6.5-报关单\BFPeaky-US-{fc}-报关资料-0605.xlsx")
    pair("BF", f"投保-{fc}",
         rf"{BF_OUT}\{fc}\BFPeaky-投保-06-05-{fc}.xlsx",
         rf"{BF_SRC}\BFPeaky-US-6.5投保\BFPeaky-投保-06-05-{fc}.xlsx")

# HUHOLE-US-6.12（加测：报关+采购合同；投保因成品缺货代单号本系统按设计跳过，
# 托书为另一变体模板不在本轮范围；投保仓库货值走 FC 配对特殊比对 diff_cangku）
HU12_SRC = CASE + r"\HUHOLE-US-6.12"
HU12_OUT = OUT + r"\HUHOLE-US-0612K"
for fc in ["GYR2", "LAS1", "LGB8", "MQJ1", "TOL3"]:
    pair("HU12", f"报关-{fc}",
         rf"{HU12_OUT}\{fc}\HUHOLE-US-{fc}-报关资料-0612.xlsx",
         rf"{HU12_SRC}\HUHOLE-US-6.12-报关单\HUHOLE-US-{fc}-报关资料-0612.xlsx")
pair("HU12", "采购合同",
     rf"{HU12_OUT}\HUHOLE-采购合同-0612.xlsx",
     rf"{HU12_SRC}\HUHOLE-US-6.12-报关单\HUHOLE-采购合同-0612.xlsx")

# HUHOLE-US-6.5 盈和（盈和精验票：成品仅托书）
HUY_OUT = OUT + r"\HUHOLE-US-0605Y"
for fc in ["ABE8", "GYR2", "IND9", "LAS1", "SBD1"]:
    pair("HUY", f"盈和托书-{fc}",
         rf"{HUY_OUT}\{fc}\HUHOLE-普船-{fc}-06.05-托书.xlsx",
         rf"{HU_SRC}\HUHOLE-US-6.5\HUHOLE-US-6.5盈和\HUHOLE-US-6.5托书1"
         rf"\HUHOLE-普船-{fc}-06.05-托书.xlsx")

# ---------------------------------------------------------------- 定性规则
# (case, 标签前缀, sheet, 单元格正则) → (分类, 说明)
RULES = [
    ("SE", "9810", "订单导入模板", r"^B\d+$", "日期类差异",
     "9810报送时间=生成时刻（RPA跑批 20260612093500，我方为本次生成时间戳）"),
    ("SE", "投保", "批量导入投保数据", r"^U3$", "日期类差异",
     "起运日期：RPA写跑批当天的周五(2026-06-12)，我方写发货基准周五(2026/06/19)"),
    ("HUK", "投保", "批量导入投保数据", r"^U3$", "RPA bug",
     "RPA起运日期(必填项)写入未生效成品为空；我方按基准周五补全（同P2第3类bug）"),
    ("BF", "投保", "批量导入投保数据", r"^U3$", "RPA bug",
     "RPA起运日期(必填项)写入未生效成品为空；我方按基准周五补全（同P2第3类bug）"),
    # ---- Serenorch 商品主数据漂移（RPA 跑批用的商品列表 ≠ 案例文件夹随附版本；
    # 与 P2 同一组字段：英文报关名/海关编码/分类/FT-1箱规高。本系统按现行主数据
    # （=案例文件夹 6.19 版商品列表）生成 ----
    ("SE", "托书", "模板", r"^F18$", "数据缺口",
     "FT-1箱规高：主数据现值77cm(=30.31in)，RPA跑批master为44cm(=17.32in)（同P2）"),
    ("SE", "托书", "模板", r"^H1[89]$", "数据缺口",
     "英文报关名漂移：现值 'disposable face towel 1pc 50count'/'compressed face towel 30pc'，"
     "RPA master 为 'Disposable face towels 50pc EF'/'compressed towel 30pc'"),
    ("SE", "托书", "模板", r"^L1[89]$", "数据缺口",
     "海关编码漂移：现值6302930010，RPA master为6302930090（同P2）"),
    ("SE", "托书", "模板", r"^M1[89]$", "数据缺口",
     "分类(用途)漂移：现值'未分类'，RPA master为'一次性毛巾'（同P2）"),
    ("SE", "报关", "报关单", r"^B(20|23)$", "数据缺口",
     "海关编码漂移：现值6302930010，RPA master为6302930090（同P2）"),
]

_NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")


def norm(v):
    if isinstance(v, _dt.datetime):
        return ("date", v.date()) if (v.hour, v.minute, v.second) == (0, 0, 0) \
            else ("num", v.timestamp())
    if isinstance(v, _dt.date):
        return ("date", v)
    if isinstance(v, bool):
        return ("num", float(v))
    if isinstance(v, (int, float)):
        return ("num", float(v))
    if isinstance(v, str):
        s = v.strip()
        if _NUM_RE.match(s):
            return ("num", float(s))
        m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
        if m:
            return ("date", _dt.date(*map(int, m.groups())))
        return ("str", s)
    return ("str", str(v))


def equal(a, b):
    na, nb = norm(a), norm(b)
    if na == nb:
        return True
    if na[0] == nb[0] == "num":
        x, y = na[1], nb[1]
        return abs(x - y) <= max(1e-9, 1e-9 * max(abs(x), abs(y)))
    if na[0] == nb[0] == "date":
        return na[1] == nb[1]
    return False


def classify(case, label, sheet, addr):
    for c, pfx, sh, pat, cat, note in RULES:
        if c == case and label.startswith(pfx) and sheet == sh \
                and re.match(pat, addr):
            return cat, note
    # SE 报关单申报要素行（D21/D24/D27…，每明细块第2行）：RPA 写入未生效成品
    # 为空，我方按映射补全 item.declare_full —— 与 P2 第2类 RPA bug 相同
    if case == "SE" and label.startswith("报关") and sheet == "报关单":
        m = re.match(r"^D(\d+)$", addr)
        if m and (int(m.group(1)) - 21) % 3 == 0 and int(m.group(1)) >= 21:
            return ("RPA bug",
                    "报关单申报要素行 RPA 写入未生效（成品为空），我方补全 declare_full（同P2）")
    return "我方错误", ""


def cells(path):
    wb = load_workbook(path)
    out = {}
    for sn in wb.sheetnames:
        d = {}
        for row in wb[sn].iter_rows():
            for c in row:
                if c.value is not None and str(c.value).strip() != "":
                    d[c.coordinate] = c.value
        out[sn] = d
    wb.close()
    return out


def diff_pair(case, label, ours_path, theirs_path):
    rows = []
    stats = {"total": 0, "MATCH": 0, "日期类差异": 0, "数据缺口": 0,
             "RPA bug": 0, "我方错误": 0}
    if not os.path.exists(ours_path):
        rows.append((case, label, "-", "-", "(我方文件缺失)", "-", "我方错误",
                     ours_path))
        stats["我方错误"] += 1
        return stats, rows
    if not os.path.exists(theirs_path):
        rows.append((case, label, "-", "-", "(成品基准缺失)", "-", "我方错误",
                     theirs_path))
        stats["我方错误"] += 1
        return stats, rows
    a, b = cells(ours_path), cells(theirs_path)
    for sn in sorted(set(a) | set(b)):
        ca, cb = a.get(sn, {}), b.get(sn, {})
        for addr in sorted(set(ca) | set(cb),
                           key=lambda k: (int(re.sub(r"\D", "", k) or 0), k)):
            va, vb = ca.get(addr), cb.get(addr)
            stats["total"] += 1
            if equal(va, vb):
                stats["MATCH"] += 1
                continue
            cat, note = classify(case, label, sn, addr)
            stats[cat] += 1
            rows.append((case, label, sn, addr, repr(va)[:60], repr(vb)[:60],
                         cat, note))
    return stats, rows


def diff_cangku():
    """投保仓库货值：成品行序为 RPA 任意序，我方按 FC 字典序——按 FC 配对比值。
    返回 (stats, rows)。"""
    ours_p = rf"{HU12_OUT}\HUHOLE-US-投保仓库货值-0612.xlsx"
    theirs_p = (rf"{HU12_SRC}\HUHOLE-US-6.12-投保"
                rf"\HUHOLE-US-投保仓库货值-0612.xlsx")
    stats = {"total": 0, "MATCH": 0, "日期类差异": 0, "数据缺口": 0,
             "RPA bug": 0, "我方错误": 0}
    rows = []
    if not os.path.exists(ours_p):
        stats["我方错误"] += 1
        rows.append(("HU12", "投保仓库货值", "-", "-", "(我方文件缺失)", "-",
                     "我方错误", ours_p))
        return stats, rows

    def by_fc(path):
        wb = load_workbook(path)
        ws = wb["Sheet1"]
        head = ws["A2"].value
        d = {}
        r = 2
        while ws[f"B{r}"].value is not None:
            d[str(ws[f"B{r}"].value)] = (ws[f"C{r}"].value, ws[f"D{r}"].value)
            r += 1
        wb.close()
        return head, d

    ha, da = by_fc(ours_p)
    hb, dbv = by_fc(theirs_p)
    for key, va, vb in [("A2", ha, hb)]:
        stats["total"] += 1
        if equal(va, vb):
            stats["MATCH"] += 1
        else:
            stats["我方错误"] += 1
            rows.append(("HU12", "投保仓库货值", "Sheet1", key,
                         repr(va), repr(vb), "我方错误", ""))
    for fc in sorted(set(da) | set(dbv)):
        for i, col in enumerate(("C", "D")):
            va = da.get(fc, (None, None))[i]
            vb = dbv.get(fc, (None, None))[i]
            stats["total"] += 1
            if equal(va, vb):
                stats["MATCH"] += 1
            else:
                stats["我方错误"] += 1
                rows.append(("HU12", "投保仓库货值", "Sheet1", f"{fc}.{col}",
                             repr(va), repr(vb), "我方错误", ""))
    flag = "OK " if stats["我方错误"] == 0 else "FAIL"
    print(f"[{flag}] HU12/投保仓库货值(按FC配对，行序差异见报告): "
          f"{stats['total']}格 M{stats['MATCH']} 错{stats['我方错误']}")
    return stats, rows


def main():
    want = set(sys.argv[1:]) or {"SE", "HUK", "BF", "HUY", "HU12"}
    agg = {}
    all_rows = []
    for case, label, ours, theirs in PAIRS:
        if case not in want:
            continue
        stats, rows = diff_pair(case, label, ours, theirs)
        all_rows.extend(rows)
        flag = "OK " if stats["我方错误"] == 0 else "FAIL"
        print(f"[{flag}] {case}/{label}: {stats['total']}格 "
              f"M{stats['MATCH']} 日{stats['日期类差异']} 缺{stats['数据缺口']} "
              f"R{stats['RPA bug']} 错{stats['我方错误']}")
        k = case
        agg.setdefault(k, dict.fromkeys(stats, 0))
        for kk, v in stats.items():
            agg[k][kk] += v
    if "HU12" in want and os.path.exists(HU12_OUT):
        stats, rows = diff_cangku()
        all_rows.extend(rows)
        agg.setdefault("HU12", dict.fromkeys(stats, 0))
        for kk, v in stats.items():
            agg["HU12"][kk] += v
    print()
    for k, s in agg.items():
        print(f"== {k}: total {s['total']} MATCH {s['MATCH']} "
              f"日期 {s['日期类差异']} 缺口 {s['数据缺口']} RPAbug {s['RPA bug']} "
              f"我方错误 {s['我方错误']}")
    print()
    n = 0
    for r in all_rows:
        if r[6] == "我方错误":
            n += 1
            print("ERR", r[0], r[1], r[2], r[3], "| ours=", r[4], "| rpa=", r[5])
            if n > 120:
                print("...(更多省略)")
                break


if __name__ == "__main__":
    main()

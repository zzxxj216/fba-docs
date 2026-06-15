"""Excel 导入：产品（"商品列表-报关"表头）/ 品牌-工厂-主体（店铺主体-工厂信息库.xlsx）
/ 离线历史批次（赛狐"发货单导出"xlsx，API 拉不到的老批次用它重建）。

- 产品按 SKU upsert；导入值为空(None/空串)时不覆盖已有值
- 工厂按名字去重、主体按名字去重、品牌按名字 upsert 并绑定工厂+主体
- "星盟" sheet → Company(type=trade)
- 离线批次 inbound_plan_id = "offline-{文件名hash前12}"，幂等重导
"""

import hashlib
import io
import json
import os
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from ..models import Batch, Brand, Company, Factory, Product, Shipment, ShipmentItem
from . import validate_service
from .sync_service import _set, next_friday, prev_month_same_day

# 产品导入模板的完整表头（与"商品列表-报关"一致，可直接回填）
PRODUCT_TEMPLATE_HEADERS = [
    "SKU", "品名", "英文品名", "分类", "商品品牌", "材质", "采购成本(￥)",
    "报关单价(US$)", "箱规长(cm)", "箱规宽(cm)", "箱规高(cm)", "单箱重量(kg)",
    "单箱数量(pcs)", "商品包装规格长(cm)", "商品包装规格宽(cm)", "商品包装规格高(cm)",
    "商品包装重量", "商品包装重量单位", "中文报关名（报关单）", "英文报关名",
    "海关编码", "中文品名（合同）", "货物名称（发票箱单）", "中文品名(用途功能）",
]

# 归一化表头 → Product 字段
PRODUCT_HEADER_MAP = {
    "SKU": "sku",
    "中文报关名(报关单)": "name_customs_cn",
    "英文报关名": "name_customs_en",
    "中文品名(合同)": "name_contract",
    "货物名称(发票箱单)": "name_invoice",
    "中文品名(用途功能)": "name_usage",
    "海关编码": "hs_code",
    "申报要素": "declare_elements",
    "材质": "material",
    "分类": "usage",
    "商品品牌": "brand_name",
    "型号": "model",
    "采购成本(￥)": "purchase_cost_default",
    "报关单价(US$)": "unit_price_default",
    "单箱重量(kg)": "box_weight_kg",
    "箱规长(cm)": "carton_l_cm",
    "箱规宽(cm)": "carton_w_cm",
    "箱规高(cm)": "carton_h_cm",
    "单箱数量(pcs)": "qty_per_box",
}
FLOAT_FIELDS = {"purchase_cost_default", "unit_price_default", "box_weight_kg",
                "carton_l_cm", "carton_w_cm", "carton_h_cm"}
INT_FIELDS = {"qty_per_box"}


def _norm(h):
    """表头归一化：去空白、全角括号/货币符 → 半角。"""
    if h is None:
        return ""
    s = str(h).strip()
    for a, b in (("（", "("), ("）", ")"), ("＄", "$"), ("　", ""), (" ", "")):
        s = s.replace(a, b)
    return s


def _load_wb(file):
    """file 可以是路径(str)或 bytes。data_only=True 取公式计算结果。

    read_only 模式下部分赛狐导出的 dimension 元数据损坏（声称只有 1 行/1 列），
    逐 sheet reset_dimensions 让 openpyxl 按真实内容迭代。"""
    if isinstance(file, (bytes, bytearray)):
        file = io.BytesIO(file)
    wb = load_workbook(file, data_only=True, read_only=True)
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()
        except AttributeError:
            pass
    return wb


def _cell_str(v):
    return "" if v is None else str(v).strip()


def _cell_float(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None       # 公式无缓存值/非数字 → 不导入该格


# ---------------------------------------------------------------- 产品

def import_products(db, file):
    """兼容"商品列表-报关"表头，按 SKU upsert → {"created", "updated"}。"""
    wb = _load_wb(file)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {"created": 0, "updated": 0}

    col_map = {}
    for idx, h in enumerate(header):
        field = PRODUCT_HEADER_MAP.get(_norm(h))
        if field:
            col_map[idx] = field
    if "sku" not in col_map.values():
        raise ValueError("导入文件缺少 SKU 列（表头需兼容'商品列表-报关'格式）")

    created = updated = 0
    for row in rows:
        data = {}
        for idx, field in col_map.items():
            v = row[idx] if idx < len(row) else None
            if field in FLOAT_FIELDS:
                data[field] = _cell_float(v)
            elif field in INT_FIELDS:
                data[field] = _cell_int(v)
            else:
                data[field] = _cell_str(v)
        sku = data.pop("sku", "")
        if not sku:
            continue
        p = db.query(Product).filter(Product.sku == sku).first()
        if p is None:
            p = Product(sku=sku)
            db.add(p)
            created += 1
        else:
            updated += 1
        for field, v in data.items():
            if v not in (None, ""):          # 空值不覆盖已有数据
                setattr(p, field, v)
        # 行序：最近一次导入文件的行序（采购合同聚合按此排，见 field_registry）
        p.sort_index = created + updated
    db.commit()
    wb.close()
    return {"created": created, "updated": updated}


def product_import_template():
    """产品导入空模板（bytes）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品"
    ws.append(PRODUCT_TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 品牌/工厂/主体

def _row_dict(header, row):
    return {_norm(h): (row[i] if i < len(row) else None)
            for i, h in enumerate(header) if h}


def import_brands(db, file):
    """"店铺主体-工厂信息库.xlsx"：品牌-工厂 sheet + 星盟 sheet。"""
    wb = _load_wb(file)
    result = {"factories_created": 0, "companies_created": 0,
              "brands_created": 0, "brands_updated": 0, "trade_created": 0}

    # ---- 品牌-工厂 sheet（找名字含"品牌"的，找不到用第一个）----
    sheet_bf = next((s for s in wb.sheetnames if "品牌" in s), wb.sheetnames[0])
    ws = wb[sheet_bf]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or []
    for row in rows:
        r = _row_dict(header, row)
        brand_name = _cell_str(r.get("品牌"))
        if not brand_name:
            continue

        # 工厂按名字去重
        factory = None
        fname = _cell_str(r.get("工厂名字"))
        if fname:
            factory = db.query(Factory).filter(Factory.name == fname).first()
            if factory is None:
                factory = Factory(name=fname,
                                  abbr=_cell_str(r.get("工厂字母缩写")),
                                  address=_cell_str(r.get("工厂地址")),
                                  phone=_cell_str(r.get("工厂电话")))
                db.add(factory)
                db.flush()
                result["factories_created"] += 1

        # 店铺主体按名字去重
        company = None
        cname = _cell_str(r.get("店铺主体"))
        if cname:
            company = db.query(Company).filter(Company.name_cn == cname).first()
            if company is None:
                company = Company(type="shop", name_cn=cname,
                                  address_cn=_cell_str(r.get("店铺地址")),
                                  phone=_cell_str(r.get("店铺电话")),
                                  uscc=_cell_str(r.get("信用代码")),
                                  name_en=_cell_str(r.get("公司名英文")),
                                  address_en=_cell_str(r.get("公司地址英文")))
                db.add(company)
                db.flush()
                result["companies_created"] += 1

        # 品牌按名字 upsert + 绑定
        brand = db.query(Brand).filter(Brand.name == brand_name).first()
        if brand is None:
            brand = Brand(name=brand_name, abbr2=brand_name[:2].upper())
            db.add(brand)
            result["brands_created"] += 1
        else:
            result["brands_updated"] += 1
        if factory:
            brand.factory_id = factory.id
        if company:
            brand.company_id = company.id
        if not brand.abbr2:
            brand.abbr2 = brand_name[:2].upper()

    # ---- 星盟 sheet → Company(type=trade) ----
    sheet_sa = next((s for s in wb.sheetnames if "星盟" in s), None)
    if sheet_sa:
        ws = wb[sheet_sa]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or []
        for row in rows:
            r = _row_dict(header, row)
            name = _cell_str(r.get("星盟主体"))
            if not name:
                continue
            company = db.query(Company).filter(Company.name_cn == name).first()
            if company is None:
                db.add(Company(type="trade", name_cn=name,
                               address_cn=_cell_str(r.get("星盟地址")),
                               phone=_cell_str(r.get("电话")),
                               abbr=_cell_str(r.get("缩写")) or "SA",
                               name_en=_cell_str(r.get("公司名英文")),
                               address_en=_cell_str(r.get("公司地址英文"))))
                result["trade_created"] += 1

    db.commit()
    wb.close()
    return result


# ---------------------------------------------------------------- 离线历史批次

SHIP_DETAIL_SHEET = "发货单详情"
PACKING_SHEET = "装箱信息"
# 新版赛狐「FBA货件导出」（HUBFSE 案例数据）：货件级地址 + 逐(货件×SKU)装箱行
FBA_DETAIL_SHEET = "货件详情"
FBA_PACK_SHEET = "装箱明细"


def _cell_int(v):
    try:
        return int(float(v)) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _base_header(h):
    """去括号后缀并小写：'Reference ID(发货商品)' → 'referenceid'。"""
    return _norm(h).split("(")[0].lower()


def _sheet_dicts(ws):
    """首行为表头，逐行产出 {基础表头: 原始值}，跳过全空行。"""
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or []
    keys = [_base_header(h) for h in header]
    for row in rows:
        if row is None or all(v in (None, "") for v in row):
            continue
        yield {k: (row[i] if i < len(row) else None)
               for i, k in enumerate(keys) if k}


def import_shipping_order(db, file, filename="", batch_name=""):
    """赛狐「发货单导出」xlsx → 重建离线历史批次 → {"batch_id","created","report"}。

    - sheet「发货单详情」每行一个(货件×SKU)：店铺/配送地址(FC)/MSKU/申报量/Reference ID
    - sheet「装箱信息」货件级：货件编号(FBA号)/箱数，按 ReferenceId 关联详情行
    - Batch 按 inbound_plan_id="offline-{文件名hash前12}" 幂等；
      Shipment 按 reference_id、Item 按 msku 增量更新，edited_fields 保护同 sync
    - 离线数据没有配送地址明细 → 地址留空，Shipment.remark 标记 "offline"
    - 箱数是货件级的：单 SKU 货件 box_count=货件箱数；多 SKU 货件无法拆分 → 留空
    """
    # ---- 幂等键：文件名 hash（无文件名时退化为内容 hash）----
    if not filename and isinstance(file, str):
        filename = os.path.basename(file)
    if filename:
        digest = hashlib.sha256(os.path.basename(filename).encode("utf-8")).hexdigest()
    else:
        content = bytes(file) if isinstance(file, (bytes, bytearray)) else \
            open(file, "rb").read()
        digest = hashlib.sha256(content).hexdigest()
    inbound_plan_id = "offline-" + digest[:12]

    # ---- 阶段1：解析 Excel（失败不动库）----
    wb = _load_wb(file)
    if FBA_DETAIL_SHEET in wb.sheetnames:        # 新版「FBA货件」导出
        return _import_fba_export(db, wb, inbound_plan_id, batch_name)
    if SHIP_DETAIL_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"文件缺少「{SHIP_DETAIL_SHEET}」sheet，不是赛狐发货单导出文件")

    details = {}                       # ref -> [{"fc","msku","qty"}]
    shop_name = ""
    for r in _sheet_dicts(wb[SHIP_DETAIL_SHEET]):
        ref = _cell_str(r.get("referenceid"))
        msku = _cell_str(r.get("msku"))
        if not ref and not msku:
            continue
        shop_name = shop_name or _cell_str(r.get("店铺"))
        details.setdefault(ref, []).append({
            "fc": _cell_str(r.get("配送地址")),
            "msku": msku,
            "qty": _cell_int(r.get("申报量")),
        })
    if not details:
        wb.close()
        raise ValueError(f"「{SHIP_DETAIL_SHEET}」sheet 没有数据行")

    # 装箱信息：同一货件(ref)可有多行（每行=一组 SKU 的箱数，行序与
    # 「发货单详情」该 ref 的明细行一一对应——6.19 案例核实：ABE8 装箱行
    # 13/2 对应 FT-1=13箱、CT-30=2箱，与成品托书 R 列一致）。
    packing = {}                       # ref -> {"fba": str, "boxes": [int,...]}
    pack_order = []
    if PACKING_SHEET in wb.sheetnames:
        for r in _sheet_dicts(wb[PACKING_SHEET]):
            fba = _cell_str(r.get("货件编号"))
            ref = _cell_str(r.get("referenceid"))
            if not (fba or ref):
                continue
            rec = packing.get(ref)
            if rec is None:
                rec = packing[ref] = {"fba": fba, "boxes": []}
                pack_order.append(ref)
            rec["fba"] = rec["fba"] or fba
            rec["boxes"].append(_cell_int(r.get("箱数")))
    wb.close()

    # 详情有、装箱缺的 ref 也建货件（无 FBA 号/箱数）
    for ref in details:
        if ref not in packing:
            packing[ref] = {"fba": "", "boxes": []}
            pack_order.append(ref)

    # ---- 阶段2：批次 upsert ----
    batch = db.query(Batch).filter(Batch.inbound_plan_id == inbound_plan_id).first()
    created = batch is None
    if created:
        batch = Batch(inbound_plan_id=inbound_plan_id)
        db.add(batch)

    batch.shop_name = shop_name or batch.shop_name
    # 国家：店铺名尾段是 2 位字母按国家代码用（Serenorch-US → US），否则保持默认
    tail = shop_name.rsplit("-", 1)[-1].strip().upper() if "-" in shop_name else ""
    if len(tail) == 2 and tail.isalpha():
        batch.country = tail
        batch.marketplace = tail

    # 品牌按 shop_name '-' 前段忽略大小写匹配，带出主体/工厂（已绑定的不覆盖）
    prefix = (shop_name.split("-")[0] if shop_name else "").strip()
    brand = None
    if prefix:
        brand = (db.query(Brand)
                 .filter(Brand.name.ilike(prefix), Brand.active.is_(True)).first())
    if brand and not batch.brand_id:
        batch.brand_id = brand.id
        if not batch.company_id:
            batch.company_id = brand.company_id
        if not batch.factory_id:
            batch.factory_id = brand.factory_id

    # 日期规则同 sync_service：base=下一个周五、contract=上月同日；已有值不覆盖
    if not batch.base_date:
        batch.base_date = next_friday().isoformat()
    if not batch.contract_date:
        batch.contract_date = prev_month_same_day(
            date.fromisoformat(batch.base_date)).isoformat()

    if batch_name:
        batch.name = batch_name
    elif not batch.name:
        base = date.fromisoformat(batch.base_date)
        batch.name = f"{shop_name or '离线批次'}-{base.strftime('%m%d')}"
    batch.synced_at = datetime.now()
    db.flush()

    # ---- 阶段3：货件/明细 upsert（按 reference_id / msku 幂等）----
    existing_sps = {sp.reference_id: sp for sp in batch.shipments}
    incoming_refs = set()
    for ref in pack_order:
        rec = packing[ref]
        fba = rec["fba"]
        boxes = rec["boxes"]
        carton_num = sum(b for b in boxes if b) or None
        rows = details.get(ref, [])
        incoming_refs.add(ref)

        sp = existing_sps.get(ref)
        if sp is None:
            sp = Shipment(batch_id=batch.id, reference_id=ref)
            db.add(sp)
            existing_sps[ref] = sp
        edited = set(json.loads(sp.edited_fields or "[]"))

        fc_code = next((r["fc"] for r in rows if r["fc"]), "")
        _set(sp, "amazon_shipment_id", fba, edited)
        _set(sp, "fc_code", fc_code, edited)
        _set(sp, "carton_num", carton_num, edited)
        _set(sp, "country_code", batch.country or "US", edited)
        # 离线标记：地址等缺失字段供校验/下游识别降级
        if "offline" not in (sp.remark or ""):
            sp.remark = (sp.remark + ";offline") if sp.remark else "offline"
        db.flush()

        # 明细：保持详情行序；装箱行序与详情行序一一对应 → 逐 SKU 箱数；
        # 同 ref 同 SKU 多行 → 数量/箱数累加
        merged = {}                              # msku -> {"qty", "boxes"}
        for i, r in enumerate(rows):
            if not r["msku"]:
                continue
            m = merged.setdefault(r["msku"], {"qty": 0, "boxes": 0})
            m["qty"] += r["qty"] or 0
            if i < len(boxes) and boxes[i]:
                m["boxes"] += boxes[i]
        # 装箱行数与详情行数不一致（仅货件级一行箱数等）→ 退化：
        # 单 SKU 货件箱数=货件箱数，多 SKU 留空（与旧逻辑一致）
        aligned = len(boxes) == len([r for r in rows if r["msku"]])
        single_sku = len(merged) == 1

        existing_items = {it.msku: it for it in sp.items}
        for msku, m in merged.items():
            qty = m["qty"] or None
            item = existing_items.get(msku)
            if item is None:
                item = ShipmentItem(shipment_id=sp.id, msku=msku)
                db.add(item)
            iedited = set(json.loads(item.edited_fields or "[]"))
            product = db.query(Product).filter(Product.sku == msku).first()
            _set(item, "product_id", product.id if product else None, iedited)
            _set(item, "qty", qty, iedited)
            if aligned and m["boxes"]:
                bc = m["boxes"]
            elif single_sku and carton_num:
                bc = carton_num
            else:
                bc = None
            _set(item, "box_count", bc, iedited)
            per = (qty // bc) if (bc and qty and qty % bc == 0) else None
            _set(item, "qty_per_box", per, iedited)
            # 箱规 cm：明细不填，取值时由 field_registry 从 Product 的
            # carton_l_cm/carton_w_cm/carton_h_cm/qty_per_box 兜底
            if product:
                _set(item, "customs_unit_price", product.unit_price_default, iedited)
                _set(item, "purchase_cost", product.purchase_cost_default, iedited)
        # 源数据里已消失的明细删掉（同 sync 的保护：源为空不清库）
        for msku, item in existing_items.items():
            if msku not in merged and merged:
                db.delete(item)
        # 逐箱数据：发货单导出无逐箱明细 → boxes 留空

    # 文件里已不存在的货件删除
    for sp in list(batch.shipments):
        if sp.reference_id not in incoming_refs:
            db.delete(sp)

    db.commit()
    db.expire_all()

    report = validate_service.validate_batch(db, batch.id)
    return {"batch_id": batch.id, "created": created, "report": report}


# ------------------------------------------------ 新版「FBA货件」导出导入

def _import_fba_export(db, wb, inbound_plan_id, batch_name=""):
    """新版赛狐「FBA货件-xxxx.xlsx」导出（sheet：货件详情/装箱明细）。

    - 货件详情：每货件一行——货件单号(FBA)/店铺/Reference ID/物流中心编码(FC)
      /收件邮编/州/城市/街道地址（**自带完整收件地址**，无需地址库回填）
    - 装箱明细：每(货件×SKU)一行——MSKU/FNSKU/箱子毛重(LB)/箱规长宽高(IN)
      /单箱数量/箱数；首行附货件级总箱数。明细 qty = 单箱数量×箱数。
      箱规按 in×2.54 存 cm（field_registry 取 in 时 ÷2.54 round2 还原）。
    - 幂等：Shipment 按 reference_id、Item 按 msku；edited_fields 保护同 sync。
    """
    if FBA_PACK_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"文件缺少「{FBA_PACK_SHEET}」sheet，不是新版 FBA货件导出文件")

    details = []                  # 货件级
    shop_name = ""
    for r in _sheet_dicts(wb[FBA_DETAIL_SHEET]):
        fba = _cell_str(r.get("货件单号"))
        if not fba:
            continue
        shop_name = shop_name or _cell_str(r.get("店铺"))
        details.append({
            "fba": fba,
            "ref": _cell_str(r.get("referenceid") or r.get("reference id")),
            "fc": _cell_str(r.get("物流中心编码")),
            "postal": _cell_str(r.get("收件邮编")),
            "state": _cell_str(r.get("收件州/省")),
            "city": _cell_str(r.get("收件城市")),
            "street": _cell_str(r.get("收件街道地址")),
        })
    if not details:
        wb.close()
        raise ValueError(f"「{FBA_DETAIL_SHEET}」sheet 没有数据行")

    packs = {}                    # fba -> {"carton_num", "items": [..]}
    cur = None
    for r in _sheet_dicts(wb[FBA_PACK_SHEET]):
        fba = _cell_str(r.get("货件单号")) or cur
        cur = fba
        if not fba:
            continue
        rec = packs.setdefault(fba, {"carton_num": None, "items": []})
        if rec["carton_num"] is None:
            rec["carton_num"] = _cell_int(r.get("总箱数"))
        msku = _cell_str(r.get("msku"))
        if not msku:
            continue
        per_box = _cell_int(r.get("单箱数量"))
        box_count = _cell_int(r.get("箱数"))
        rec["items"].append({
            "msku": msku,
            "fnsku": _cell_str(r.get("fnsku")),
            "per_box": per_box,
            "box_count": box_count,
            "qty": (per_box * box_count) if per_box and box_count else None,
            # 箱规 IN → 存 cm（×2.54），field_registry ÷2.54 round2 可精确还原
            "l_cm": _cell_float(r.get("箱子长度")),
            "w_cm": _cell_float(r.get("箱子宽度")),
            "h_cm": _cell_float(r.get("箱子高度")),
        })
    wb.close()
    for rec in packs.values():
        for it in rec["items"]:
            for k in ("l_cm", "w_cm", "h_cm"):
                if it[k] is not None:
                    it[k] = round(it[k] * 2.54, 4)

    # ---- 批次 upsert（与旧格式同口径）----
    batch = db.query(Batch).filter(Batch.inbound_plan_id == inbound_plan_id).first()
    created = batch is None
    if created:
        batch = Batch(inbound_plan_id=inbound_plan_id)
        db.add(batch)
    batch.shop_name = shop_name or batch.shop_name
    tail = shop_name.rsplit("-", 1)[-1].strip().upper() if "-" in shop_name else ""
    if len(tail) == 2 and tail.isalpha():
        batch.country = tail
        batch.marketplace = tail
    prefix = (shop_name.split("-")[0] if shop_name else "").strip()
    brand = None
    if prefix:
        brand = (db.query(Brand)
                 .filter(Brand.name.ilike(prefix), Brand.active.is_(True)).first())
    if brand and not batch.brand_id:
        batch.brand_id = brand.id
        if not batch.company_id:
            batch.company_id = brand.company_id
        if not batch.factory_id:
            batch.factory_id = brand.factory_id
    if not batch.base_date:
        batch.base_date = next_friday().isoformat()
    if not batch.contract_date:
        batch.contract_date = prev_month_same_day(
            date.fromisoformat(batch.base_date)).isoformat()
    if batch_name:
        batch.name = batch_name
    elif not batch.name:
        base = date.fromisoformat(batch.base_date)
        batch.name = f"{shop_name or '离线批次'}-{base.strftime('%m%d')}"
    batch.synced_at = datetime.now()
    db.flush()

    # ---- 货件/明细 upsert ----
    existing_sps = {sp.reference_id: sp for sp in batch.shipments}
    incoming_refs = set()
    for d in details:
        ref = d["ref"]
        incoming_refs.add(ref)
        sp = existing_sps.get(ref)
        if sp is None:
            sp = Shipment(batch_id=batch.id, reference_id=ref)
            db.add(sp)
            existing_sps[ref] = sp
        edited = set(json.loads(sp.edited_fields or "[]"))
        pack = packs.get(d["fba"], {"carton_num": None, "items": []})
        _set(sp, "amazon_shipment_id", d["fba"], edited)
        _set(sp, "fc_code", d["fc"], edited)
        _set(sp, "address_line1", d["street"], edited)
        _set(sp, "city", d["city"], edited)
        _set(sp, "state", d["state"], edited)
        _set(sp, "postal_code", d["postal"], edited)
        _set(sp, "carton_num", pack["carton_num"], edited)
        _set(sp, "country_code", batch.country or "US", edited)
        if "offline" not in (sp.remark or ""):
            sp.remark = (sp.remark + ";offline") if sp.remark else "offline"
        db.flush()

        existing_items = {it.msku: it for it in sp.items}
        seen = set()
        for row in pack["items"]:
            msku = row["msku"]
            if msku in seen:        # 同 SKU 多行：数量/箱数累加
                item = existing_items[msku]
                iedited = set(json.loads(item.edited_fields or "[]"))
                _set(item, "qty", (item.qty or 0) + (row["qty"] or 0), iedited)
                _set(item, "box_count",
                     (item.box_count or 0) + (row["box_count"] or 0), iedited)
                continue
            seen.add(msku)
            item = existing_items.get(msku)
            if item is None:
                item = ShipmentItem(shipment_id=sp.id, msku=msku)
                db.add(item)
                existing_items[msku] = item
            iedited = set(json.loads(item.edited_fields or "[]"))
            product = db.query(Product).filter(Product.sku == msku).first()
            _set(item, "product_id", product.id if product else None, iedited)
            _set(item, "fnsku", row["fnsku"], iedited)
            _set(item, "qty", row["qty"], iedited)
            _set(item, "box_count", row["box_count"], iedited)
            _set(item, "qty_per_box", row["per_box"], iedited)
            _set(item, "carton_l", row["l_cm"], iedited)
            _set(item, "carton_w", row["w_cm"], iedited)
            _set(item, "carton_h", row["h_cm"], iedited)
            if product:
                _set(item, "customs_unit_price", product.unit_price_default, iedited)
                _set(item, "purchase_cost", product.purchase_cost_default, iedited)
        for msku, item in existing_items.items():
            if msku not in seen and seen:
                db.delete(item)

    for sp in list(batch.shipments):
        if sp.reference_id not in incoming_refs:
            db.delete(sp)

    db.commit()
    db.expire_all()
    report = validate_service.validate_batch(db, batch.id)
    return {"batch_id": batch.id, "created": created, "report": report}

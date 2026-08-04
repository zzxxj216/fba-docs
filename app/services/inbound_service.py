"""建仓输入准备服务（2026-07-29 起建仓执行改走 mcapi 的赛狐建仓接口 /api/v1/sellfox/inbound/*）。

本模块只保留建仓的**输入准备**能力，供 Codex/前端在调 mcapi 建仓前取数：
- parse_replenishment_excel：补仓计划 Excel → 明细行（列关键词匹配，长宽高 in / 重 lb 直读）
- items_from_purchase_plan：赛狐采购计划 → 明细行
- _resolve_items / _fill_box_spec_from_sellfox：每箱数/箱规缺省补全（产品库缺则回查赛狐并缓存）

注意口径：赛狐建仓的 box_specs 用 **cm/kg**（产品库原生口径）；本模块 Excel 解析出的
l_in/w_in/h_in/weight_lb 是英寸/磅（模板历史格式），喂赛狐建仓前需换算回 cm/kg。
建仓过程记录（断点续跑）见 routers/inbound.py 的 records 接口（InboundPlan 表复用）。
亚马逊 SP-API 建仓执行线已整体移除（历史实现见 git 历史）。
"""

import io
import json
import math
import os
import re

from openpyxl import load_workbook

from ..models import Brand, InboundPlan, Product

CM_PER_IN = 2.54
LB_PER_KG = 2.20462

# 发货地址：存店铺档案 Brand.source_address(JSON, mcapi FbaSourceAddress snake_case)。
# **店铺独有信息，严禁共用/回退**——曾因回退默认地址导致 Byane 建仓用了 HUHOLE 的发货地
# (店铺串联风险)。缺失 = 建仓直接报错，绝不悄悄用别家的。


def _source_address(db, brand_id):
    """按品牌取该店铺自己的发货地址(店铺档案)。缺失=报错，不回退。"""
    b = db.get(Brand, brand_id) if brand_id else None
    addr = None
    try:
        addr = json.loads(b.source_address) if (b and b.source_address) else None
    except (ValueError, TypeError):
        addr = None
    required = (
        "name", "address_line1", "city", "state_or_province_code",
        "country_code", "postal_code", "phone_number",
    )
    missing = [k for k in required if not str((addr or {}).get(k) or "").strip()]
    if missing:
        raise RuntimeError(
            f"品牌「{b.name if b else brand_id}」未配置发货地址（店铺档案）——"
            f"缺少 {', '.join(missing)}；为防店铺串联不允许回退默认地址，"
            "请先补全该店铺发货地址再建仓")
    return addr

ST_CREATED = "计划已创建"
ST_PACKED = "装箱已确认"
ST_BOXED = "已提交箱规"
ST_PLACEMENT = "分仓方案已生成"
ST_PLACED = "分仓已确认"
ST_FAIL = "失败"


# ---------------------------------------------------------------- helpers

def _get(db, rec_id):
    rec = db.query(InboundPlan).filter(InboundPlan.id == rec_id).first()
    if rec is None:
        raise RuntimeError(f"建仓记录 {rec_id} 不存在")
    return rec


def _store(db, brand_id):
    if not brand_id:
        return ""
    b = db.query(Brand).filter(Brand.id == brand_id).first()
    return (b.amazon_store or "") if b else ""


# 欧洲站批次 → 用品牌店铺的 _eu 凭据(mcapi 按区分店铺:byane/byane_eu)
_EU_COUNTRIES = {"DE", "FR", "IT", "ES", "NL", "SE", "PL", "BE", "IE", "UK", "GB", "EU"}

# 国家 → marketplace id(建计划 destinationMarketplaces 用)
_MARKETPLACE_IDS = {
    "US": "ATVPDKIKX0DER", "CA": "A2EUQ1WTGCTBG2", "MX": "A1AM78C64UM0Y8",
    "DE": "A1PA6795UKMFR9", "FR": "A13V1IB3VIYZZH", "IT": "APJ6JRA9NG5V4",
    "ES": "A1RKKUPIHCS9HS", "UK": "A1F83G8C2ARO7P", "GB": "A1F83G8C2ARO7P",
    "NL": "A1805IZSGTT6HS", "SE": "A2NODRKZP88ZB9", "PL": "A1C3SOZRARQ6R3",
    "BE": "AMEN7PMS3EDWL", "IE": "A28R8C7NBKEWEA",
}


def _store_for_batch(db, batch):
    s = _store(db, batch.brand_id)
    country = (batch.country or batch.marketplace or "").strip().upper()
    if s and country in _EU_COUNTRIES and not s.endswith("_eu"):
        return f"{s}_eu"
    return s


def _loads(s, default):
    try:
        return json.loads(s) if s else default
    except (ValueError, TypeError):
        return default


def _dict(rec):
    return {
        "id": rec.id,
        "source_type": rec.source_type,
        "source_ref": rec.source_ref,
        "brand_id": rec.brand_id,
        "store": rec.store,
        "name": rec.name,
        "status": rec.status,
        "amazon_inbound_plan_id": rec.amazon_inbound_plan_id,
        "items": _loads(rec.items_snapshot, []),
        "placement_option_id": rec.placement_option_id,
        "shipments": _loads(rec.shipments_snapshot, {}),
        "current_operation_id": rec.current_operation_id,
        "batch_id": rec.batch_id,
        "error": rec.error or "",
        "created_at": rec.created_at.strftime("%Y-%m-%d %H:%M:%S") if rec.created_at else "",
    }


def _fill_box_spec_from_sellfox(db, msku, p):
    """本地缺箱规时，从赛狐商品库 /api/commodity/pageList.json 拉箱规回填 Product 并缓存。
    赛狐字段：cartonLength/Width/Height(cm)、cartonWeight(kg/箱)、cartonQty(每箱数)。拉不到不阻塞。"""
    from .. import sellfox_client as sf

    def _f(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _i(v):
        try:
            return int(float(v)) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    # msku(Amazon)和赛狐商品库 sku 常有差异：尾部引号(6")、站点后缀(-US/-UK…)、
    # 尾部数字批次码(Zentop 的 -016/-005)、-New 后缀。依次尝试各种去后缀变体。
    import re
    bases = [msku]
    m = re.match(r"^(.*)-(US|UK|CA|DE|FR|IT|ES|JP|AU|MX|NL|SE|PL|BE|TR|SG|AE|SA|BR|IN)$",
                 msku, re.I)
    if m:
        bases.append(m.group(1))
    more = []
    for b in bases:
        more.append(b)
        b2 = re.sub(r"-New$", "", b, flags=re.I)
        if b2 != b:
            more.append(b2)
        b3 = re.sub(r"-\d{1,3}$", "", b2)      # 去尾部数字码(Zentop: B1S-66DL-016 → B1S-66DL)
        if b3 != b2:
            more.append(b3)
    variants, seen = [], set()
    for b in more:
        for v in (b, b.rstrip('"”″\' ').strip()):
            if v and v not in seen:
                seen.add(v)
                variants.append(v)
    row = None
    for sku_try in variants:
        try:
            data = sf.call("/api/commodity/pageList.json",
                           {"pageNo": 1, "pageSize": 10, "skus": [sku_try]}) or {}
            row = next((x for x in (data.get("rows") or []) if (x.get("sku") or "").strip() == sku_try), None)
        except Exception:
            row = None
        if row:
            break
    if not row:
        return p
    if p is None:
        p = Product(sku=msku)
        db.add(p)
    p.carton_l_cm = p.carton_l_cm or _f(row.get("cartonLength"))
    p.carton_w_cm = p.carton_w_cm or _f(row.get("cartonWidth"))
    p.carton_h_cm = p.carton_h_cm or _f(row.get("cartonHeight"))
    p.box_weight_kg = p.box_weight_kg or _f(row.get("cartonWeight"))
    p.qty_per_box = p.qty_per_box or _i(row.get("cartonQty"))
    # 报关信息一并回填(缺则补)——生成托书/报关资料的校验要用
    p.name_contract = p.name_contract or (row.get("name") or "")   # 赛狐商品全名(合同品名/规格截取源)
    # 自定义字段走商品详情V2(/api/commodity/v2/getCommodityDetail.json,传商品id)：
    # 「品名的英文翻译」→ name_invoice(发票货物名称)；「材质的英文翻译」→ material_en。
    # pageList 不返回自定义字段(2026-07-07 用文档密码核实)。有值即以赛狐为准。
    if row.get("id"):
        try:
            det = sf.call("/api/commodity/v2/getCommodityDetail.json", {"id": row["id"]}) or {}
            for cf in det.get("customFieldUsingVOList") or []:
                fname = (cf.get("fieldName") or "").strip()
                val = (cf.get("formatValue") or " ".join(cf.get("values") or [])).strip()
                if not val:
                    continue
                if "品名" in fname and "英文" in fname:
                    p.name_invoice = val
                elif "材质" in fname and "英文" in fname:
                    p.material_en = val
        except Exception:
            pass                                  # 详情拉不到不阻塞,回退已有值
    p.name_customs_cn = p.name_customs_cn or (row.get("declareNameCh") or "")
    p.name_customs_en = p.name_customs_en or (row.get("declareNameEn") or "")
    p.hs_code = p.hs_code or (row.get("hsCode") or "")
    # 材质：赛狐多为英文(steel)→ 拆成 中文材质(material) + 英文材质(material_en)
    raw_mat = row.get("declareMaterial") or row.get("materialQuality") or ""
    if raw_mat and not (p.material or "").strip():
        _cn = "".join(re.findall(r"[一-鿿]+", raw_mat))
        _en = " ".join(re.findall(r"[A-Za-z]+", raw_mat)).strip()
        _en2cn = {"steel": "铁", "iron": "铁", "stainless": "不锈钢", "plastic": "塑料",
                  "aluminum": "铝", "aluminium": "铝", "zinc": "锌", "copper": "铜", "pp": "塑料"}
        _cn2en = {"铁": "Steel", "钢": "Steel", "不锈钢": "Stainless Steel", "塑料": "Plastic",
                  "铝": "Aluminum", "锌": "Zinc", "铜": "Copper"}
        if not _cn and _en:
            _cn = _en2cn.get(_en.lower().split()[0], "")
        if not _en and _cn:                    # 赛狐只给中文(钢)时反查英文——两栏必须都有
            _en = _cn2en.get(_cn, "")
        p.material = _cn or raw_mat
        if _en and not (p.material_en or "").strip():
            p.material_en = _en[:1].upper() + _en[1:]
    # 兜底：已有中文材质但缺英文(或反之)时按词典互转——托书 G/H 两栏必须都有
    _cn2en_std = {"铁": "Steel", "钢": "Steel", "不锈钢": "Stainless Steel", "塑料": "Plastic",
                  "铝": "Aluminum", "锌": "Zinc", "铜": "Copper"}
    if (p.material or "").strip() and not (p.material_en or "").strip():
        p.material_en = _cn2en_std.get(p.material.strip(), "")
    p.usage = p.usage or (row.get("declareUseTo") or row.get("useTo") or "")
    p.declare_elements = p.declare_elements or (row.get("declareElements") or "")
    p.unit_price_default = p.unit_price_default or _f(row.get("declareCharge"))
    p.purchase_cost_default = p.purchase_cost_default or _f(row.get("purchaseCost"))
    db.flush()
    return p


def _resolve_items(db, raw_items):
    """补全/校验建仓明细。

    raw_items: [{msku, quantity, units_per_box?, l_in?, w_in?, h_in?, weight_lb?, expiration?}]
    缺的从 Product 补（qty_per_box / carton_cm→IN / box_weight_kg→LB）；缺关键项报错。
    """
    out = []
    for r in raw_items or []:
        msku = (r.get("msku") or "").strip()
        qty = int(r.get("quantity") or 0)
        if not msku or qty <= 0:
            raise RuntimeError(f"明细缺 msku 或数量<=0：{r}")
        p = db.query(Product).filter(Product.sku == msku).first()
        # 本地缺箱规或缺报关信息 → 从赛狐商品库回查补全并缓存(手填了 l_in 的不查箱规)
        need_box = not (p and p.carton_l_cm and p.carton_w_cm and p.carton_h_cm
                        and p.qty_per_box and p.box_weight_kg)
        need_customs = not (p and p.hs_code and p.name_customs_cn and p.material
                            and p.unit_price_default)
        if (need_box and not r.get("l_in")) or need_customs:
            p = _fill_box_spec_from_sellfox(db, msku, p)
        upb = int(r.get("units_per_box") or (p.qty_per_box if p and p.qty_per_box else 0) or 0)
        l_in = r.get("l_in") or (round(p.carton_l_cm / CM_PER_IN, 2) if p and p.carton_l_cm else None)
        w_in = r.get("w_in") or (round(p.carton_w_cm / CM_PER_IN, 2) if p and p.carton_w_cm else None)
        h_in = r.get("h_in") or (round(p.carton_h_cm / CM_PER_IN, 2) if p and p.carton_h_cm else None)
        wt = r.get("weight_lb") or (round(p.box_weight_kg * LB_PER_KG, 2) if p and p.box_weight_kg else None)
        miss = [k for k, v in (("每箱数", upb), ("箱长", l_in), ("箱宽", w_in), ("箱高", h_in), ("箱重", wt)) if not v]
        if miss:
            raise RuntimeError(f"SKU {msku} 缺 {('、').join(miss)}（请在产品库补箱规，或建仓时手填）")
        boxes = int(r.get("boxes") or 0)
        if boxes:
            if upb * boxes != qty:
                raise RuntimeError(
                    f"SKU {msku} 申报量 {qty} ≠ 每箱数 {upb} × 箱数 {boxes}；"
                    "赛狐 STA 装箱要求数量完全一致，请先修正箱规")
        else:
            if qty % upb:
                raise RuntimeError(
                    f"SKU {msku} 申报量 {qty} 不能被每箱数 {upb} 整除；"
                    "当前赛狐建仓只支持同一 SKU 整箱装箱，请明确调整每箱数/箱数")
            boxes = qty // upb
        out.append({"msku": msku, "quantity": qty, "units_per_box": upb,
                    "boxes": boxes, "l_in": l_in, "w_in": w_in,
                    "h_in": h_in, "weight_lb": wt,
                    "expiration": (r.get("expiration") or "").strip()})
    if not out:
        raise RuntimeError("建仓明细为空")
    return out


def box_specs_from_items(items):
    """本机 in/lb 建仓快照转 MCAPI 赛狐所需 cm/kg，并执行单箱重量红线。"""
    specs = []
    for item in items or []:
        weight_kg = round(float(item["weight_lb"]) / LB_PER_KG, 3)
        if weight_kg > 22.7:
            raise RuntimeError(
                f"SKU {item['msku']} 单箱 {weight_kg}kg 超过 22.7kg，已阻止建仓")
        specs.append({
            "msku": item["msku"],
            "per_box": int(item["units_per_box"]),
            "boxes": int(item["boxes"]),
            "length": round(float(item["l_in"]) * CM_PER_IN, 2),
            "width": round(float(item["w_in"]) * CM_PER_IN, 2),
            "height": round(float(item["h_in"]) * CM_PER_IN, 2),
            "weight_kg": weight_kg,
        })
    return specs


# ---------------------------------------------------------------- 来源：补仓计划 Excel / 赛狐采购计划

# 列关键词（归一化后按子串匹配）；声明顺序=认领顺序，先认领更专指的列
_COL_KEYS = [
    ("msku", ["amazon-sku", "amazonsku", "amazon sku", "msku", "sku"]),
    ("units_per_box", ["每箱商品数", "每箱数", "每箱", "units per box", "unitsperbox", "units/box"]),
    ("boxes", ["外箱数", "箱数", "number of boxes", "numberofboxes", "外箱", "boxes"]),
    ("quantity", ["补仓数量", "补货数量", "数量", "quantity", "qty"]),
    ("l_in", ["box length", "length", "长"]),
    ("w_in", ["box width", "width", "宽"]),
    ("h_in", ["box height", "height", "高"]),
    ("weight_lb", ["box weight", "weight", "重量", "重"]),
]


def _norm(s):
    return str(s or "").strip().lower().replace(" ", "")


def _num(v):
    if v is None or v == "":
        return None
    try:
        f = float(str(v).replace(",", ""))
        return int(f) if f == int(f) else round(f, 2)
    except (ValueError, TypeError):
        return None


def parse_replenishment_excel(content):
    """解析补仓计划 Excel → 建仓明细 raw_items。

    按表头关键词映射列：AMAZON-SKU/补仓数量/每箱商品数/外箱数/长/宽/高/重(IN/LB)。
    返回 [{msku,quantity,units_per_box,boxes,l_in,w_in,h_in,weight_lb}]；缺的留空（建仓时从产品库补/手填）。
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise RuntimeError("Excel 为空")

    # 找表头行：前 10 行里关键词命中最多的一行
    best_i, best_map, best_hits = 0, {}, -1
    for i, row in enumerate(rows[:10]):
        norm_cells = [_norm(c) for c in row]
        claimed = set()
        colmap = {}
        for field, keys in _COL_KEYS:
            for ci, cell in enumerate(norm_cells):
                if ci in claimed or not cell:
                    continue
                if any(k.replace(" ", "") in cell for k in keys):
                    colmap[field] = ci
                    claimed.add(ci)
                    break
        hits = len(colmap)
        if hits > best_hits:
            best_i, best_map, best_hits = i, colmap, hits

    if "msku" not in best_map or "quantity" not in best_map:
        raise RuntimeError("未识别到 SKU / 数量列，请确认补仓计划表头（需含 AMAZON-SKU、补仓数量等）")

    items = []
    for row in rows[best_i + 1:]:
        msku = str(row[best_map["msku"]] or "").strip() if best_map.get("msku") is not None else ""
        qty = _num(row[best_map["quantity"]]) if best_map.get("quantity") is not None else None
        if not msku or not qty:
            continue
        it = {"msku": msku, "quantity": int(qty)}
        for f in ("units_per_box", "boxes", "l_in", "w_in", "h_in", "weight_lb"):
            ci = best_map.get(f)
            it[f] = _num(row[ci]) if ci is not None else None
        items.append(it)
    if not items:
        raise RuntimeError("补仓计划没有有效明细行")
    return items


def items_from_purchase_plan(db, plan_group_no):
    """从赛狐采购计划取建仓明细（SKU+采购量）；箱规留空（建仓时从产品库补）。"""
    from . import purchase_plan_service as pps
    grp = pps._find_group(db, plan_group_no)
    if grp is None:
        raise RuntimeError(f"采购计划 {plan_group_no} 未找到")
    items = []
    for it in grp.get("purchasePlanItemVoList") or []:
        msku = str(it.get("msku") or it.get("sku") or "").strip()
        qty = _num(it.get("planNum"))
        if not msku or not qty:
            continue
        items.append({"msku": msku, "quantity": int(qty),
                      "units_per_box": _num(it.get("cartonQty")),
                      "boxes": _num(it.get("cartonNum")),
                      "l_in": None, "w_in": None, "h_in": None, "weight_lb": None})
    brand_name = ""
    for it in grp.get("purchasePlanItemVoList") or []:
        if it.get("brandName"):
            brand_name = it["brandName"]
            break
    return {"items": items, "brand_name": brand_name,
            "name": _norm_plan_name(grp)}


def _norm_plan_name(grp):
    its = grp.get("purchasePlanItemVoList") or []
    shop = (its[0].get("shopName") if its else "") or ""
    site = (its[0].get("siteName") if its else "") or ""
    return f"{shop or grp.get('planGroupNo','')}-{site}".strip("-")


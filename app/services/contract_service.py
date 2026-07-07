"""两级采购合同生成（朗格系 Byane/Zentop/RazEdg）——按 docs/doc_rules/采购合同.md v0.2。

每批次两份：
  一级：工厂(朗格锯链) → 店铺主体，单价=成本×vat_factor(1.13)，号 {店2}-LG-{合同日期}，交货地点=工厂
  二级：店铺主体 → 星盟，  单价=成本×vat_factor×markup_factor(1.13×1.1)，号 SA-{店2}-{合同日期}，交货地点=货代仓库

模板：templates_store/采购合同-两级公用模板.xlsx（版式同用户五份资料，明细17行起，
尾部总值/大写/交货时间随明细行数下移）。品名=赛狐商品全名(commodityName)，
缺时从赛狐商品库拉取并缓存到 Product.name_contract。
系数走 rule_engine（vat_factor/markup_factor），不硬编码。
"""
import json
import os
import re

import openpyxl

from .. import rule_engine
from ..excel_engine import _copy_row_style
from ..models import Batch, Brand, Company, Factory, GeneratedDoc, Product

TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
                        "amazon", "templates_store", "采购合同-两级公用模板.xlsx")
if not os.path.exists(TEMPLATE):     # 兼容不同部署根
    TEMPLATE = r"D:\amazon\templates_store\采购合同-两级公用模板.xlsx"

OUTPUT_DIR = os.path.join(r"D:\amazon", "output")

# 模板固定坐标（与样本一致）
DETAIL_START = 17          # 明细首行
BASE_CAPACITY = 11         # 不插行时可容纳的明细行数(17..27)
BASE_TOTAL_ROW = 28        # 总值行（G28=SUM(G17:G27)）
MAX_COL = 12


def _factory_party(f: Factory):
    return {"name": f.name, "addr": f.address or "", "phone": f.phone or ""}


def _company_party(c: Company):
    return {"name": c.name_cn, "addr": c.address_cn or "", "phone": c.phone or ""}


def _chain(db, brand: Brand):
    """朗格系两级链：[(卖方, 买方, 系数, 合同号前缀, 交货地点, 文件名段)] ×2。"""
    factory = db.get(Factory, brand.factory_id) if brand.factory_id else None
    store_co = db.get(Company, brand.company_id) if brand.company_id else None
    trade = (db.query(Company).filter(Company.type == "trade", Company.active.is_(True))
             .first())
    if not (factory and store_co and trade):
        raise RuntimeError(f"品牌 {brand.name} 档案不全（工厂/主体/外贸主体）")
    b2 = (brand.abbr2 or brand.name[:2]).upper()
    vat = rule_engine.get_rule_float(db, "vat_factor", None, 1.13)
    markup = rule_engine.get_rule_float(db, "markup_factor", None, 1.10)
    return [
        (_factory_party(factory), _company_party(store_co), vat,
         f"{b2}-LG", "工厂", f"LG2{b2}"),
        (_company_party(store_co), _company_party(trade), vat * markup,
         f"SA-{b2}", "货代仓库", f"{b2}2SA"),
    ]


def _commodity_name(db, msku):
    """品名=赛狐商品全名。优先 Product.name_contract 缓存，缺则拉赛狐商品库回填。"""
    p = db.query(Product).filter(Product.sku == msku).first()
    if p and (p.name_contract or "").strip():
        return p.name_contract.strip()
    from .. import sellfox_client as sf
    from .inbound_service import _fill_box_spec_from_sellfox  # 复用其SKU变体链? 直接查原值+去尾码
    for sku_try in (msku, re.sub(r"-\d{1,3}$", "", msku)):
        try:
            data = sf.call("/api/commodity/pageList.json",
                           {"pageNo": 1, "pageSize": 10, "skus": [sku_try]}) or {}
            row = next((x for x in (data.get("rows") or [])
                        if (x.get("sku") or "").strip() == sku_try), None)
        except Exception:
            row = None
        name = ((row.get("name") or row.get("commodityName") or "").strip()) if row else ""
        if name:
            if p:
                p.name_contract = name
                db.flush()
            return name
    return ""


def _aggregate_items(db, batch):
    """整批 SKU 聚合：[{name, qty, cost}]，按明细行首次出现排序。"""
    agg, order = {}, []
    for sp in batch.shipments:
        for it in sp.items:
            if not it.msku or (it.qty or 0) <= 0:
                continue
            if it.msku not in agg:
                agg[it.msku] = {"msku": it.msku, "qty": 0, "cost": it.purchase_cost or 0}
                order.append(it.msku)
            agg[it.msku]["qty"] += it.qty or 0
            if not agg[it.msku]["cost"]:
                agg[it.msku]["cost"] = it.purchase_cost or 0
    rows = []
    missing = []
    for k in order:
        r = agg[k]
        name = _commodity_name(db, r["msku"])
        if not name:
            missing.append(r["msku"])
        if not r["cost"]:
            missing.append(f"{r['msku']}(无采购成本)")
        rows.append({"name": name or r["msku"], "qty": r["qty"], "cost": r["cost"]})
    return rows, missing


def _fill_contract(dst, seller, buyer, no, contract_date, deliver_date, place, rows):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active
    n = len(rows)
    shift = max(0, n - BASE_CAPACITY)
    if shift:
        # openpyxl 的 insert_rows 只移动单元格值、**不移动合并区**——须手动:
        # 拆掉插入点以下全部合并 → 插行 → 按位移重新合并
        below = [(m.min_row, m.min_col, m.max_row, m.max_col)
                 for m in list(ws.merged_cells.ranges) if m.min_row >= 24]
        for m in [m for m in list(ws.merged_cells.ranges) if m.min_row >= 24]:
            ws.unmerge_cells(str(m))
        ws.insert_rows(24, shift)          # 在明细区空行段内插行，尾部整体下移
        for r1, c1, r2, c2 in below:
            ws.merge_cells(start_row=r1 + shift, start_column=c1,
                           end_row=r2 + shift, end_column=c2)
    total_row = BASE_TOTAL_ROW + shift

    # 头部
    ws["B2"] = seller["name"]; ws["B3"] = seller["addr"]; ws["B5"] = seller["phone"]
    ws["B7"] = buyer["name"]; ws["B9"] = buyer["addr"]; ws["B11"] = buyer["phone"]
    ws["F5"] = no
    ws["F7"] = contract_date

    # 清明细区（模板自带样本值），再写入
    for r in range(DETAIL_START, total_row):
        for col in (1, 3, 4, 5, 6, 7):
            ws.cell(r, col).value = None
    merged = {str(m) for m in ws.merged_cells.ranges}
    for i, it in enumerate(rows):
        r = DETAIL_START + i
        if r > DETAIL_START:
            _copy_row_style(ws, DETAIL_START, r, MAX_COL)
        if f"A{r}:B{r}" not in merged:      # 品名列 A:B 逐行合并(模板样式;插入的新行要补)
            ws.merge_cells(f"A{r}:B{r}")
        ws.cell(r, 1, it["name"])                       # A 品名=赛狐商品全名
        ws.cell(r, 3, it["qty"])                        # C 数量(整批聚合)
        ws.cell(r, 4, "盒")                             # D 单位
        ws.cell(r, 5, round(it["price"], 2))            # E 单价=成本×系数
        ws.cell(r, 6, "人民币")                          # F
        ws.cell(r, 7, f"=E{r}*C{r}")                    # G 金额

    # 尾部动态：总值/大写/交货时间（相对总值行的固定偏移）
    ws.cell(total_row, 7).value = f"=SUM(G{DETAIL_START}:G{total_row - 1})"
    ws.cell(total_row + 2, 2).value = f"=G{total_row}"
    ws.cell(total_row + 4, 1).value = f"(7) 交货时间：{deliver_date.strftime('%Y/%m/%d')}"
    ws.cell(total_row + 4, 3).value = f"（8）交货地点：{place}"

    os.makedirs(os.path.dirname(dst), exist_ok=True)
    wb.save(dst)
    return dst


def generate_two_level(db, batch_id, persist=True):
    """给朗格系批次生成两份采购合同。返回 {generated:[路径], warnings:[...]}。"""
    batch = db.get(Batch, batch_id)
    if batch is None:
        raise RuntimeError(f"批次不存在：{batch_id}")
    brand = db.get(Brand, batch.brand_id) if batch.brand_id else None
    if brand is None:
        raise RuntimeError("批次无品牌，无法定合同链")
    base = (rule_engine._to_date(batch.base_date) if (batch.base_date or "").strip()
            else rule_engine.next_friday())                 # 基准周五(缺则现算，同 generate_service 口径)
    contract_date = rule_engine.prev_month_same_day(base)   # 合同日期=上月同日
    rows, missing = _aggregate_items(db, batch)
    if not rows:
        raise RuntimeError("批次无明细")

    out = {"generated": [], "warnings": []}
    if missing:
        out["warnings"].append("缺赛狐全名/成本：" + ", ".join(sorted(set(missing))))
    date_s = contract_date.strftime("%Y-%m-%d")
    mmdd = base.strftime("%m%d") if hasattr(base, "strftime") else str(base)
    batch_dir = os.path.join(OUTPUT_DIR, re.sub(r'[\\/:*?"<>|]', "_", batch.name or f"batch-{batch.id}"),
                             "采购合同")
    for seller, buyer, factor, prefix, place, tag in _chain(db, brand):
        crows = [{**r, "price": (r["cost"] or 0) * factor} for r in rows]
        no = f"{prefix}-{date_s}"
        fname = f"{brand.name}-采购合同-{tag}-{mmdd}.xlsx"
        dst = os.path.join(batch_dir, fname)
        _fill_contract(dst, seller, buyer, no, contract_date, base, place, crows)
        out["generated"].append(dst)
        if persist:
            db.add(GeneratedDoc(batch_id=batch.id, shipment_id=None, template_id=None,
                                doc_type="采购合同", filename=fname, path=dst))
    db.commit()
    return out

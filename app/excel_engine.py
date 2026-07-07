"""Excel 模板填充引擎（坐标 + 锚点行 + 明细区，格式见 DESIGN.md §6）。

明细区(tables)扩展：
- 可选 "row_step": N（每条记录占 N 行，默认 1），
  columns 键支持 "D+1" 语法 = 写到 start_row + i*row_step + 1 行的 D 列。
- 可选 "reserved_rows": N（模板预留了 N 条记录的格式行）。实际记录数
  少于 N 时，多余的预留行整行删除，并按 Excel 语义调整全工作簿公式
  （同 sheet 引用与跨 sheet 引用：行号下移、SUM 区间收缩；引用落在被
  删行内的单元格引用置 #REF!）。货代报关模板"合计行"公式（如
  =SUM(C8:C9)、=箱单!F10）因此能在删行后保持正确。

铁律（CLAUDE.md 约定 2）：写入永远显式取 sheet —— mapping 里 name 缺省时
取第一个 sheet，但写入必须经由确定的 ws 对象，杜绝"空 sheet 名写错工作表"。
出错抛 ValueError，带可读中文信息（哪个 sheet 哪个格子哪个路径）。
"""

import os
import re
from copy import copy

from openpyxl import load_workbook

from .field_registry import resolve

# 明细区列名："D" 或 "D+1"（=写到该记录基准行的下 1 行的 D 列，配合 row_step 用）
_COL_RE = re.compile(r"^([A-Za-z]{1,3})(?:\+(\d+))?$")


def _get_sheet(wb, spec):
    name = (spec.get("name") or "").strip()
    if not name:
        return wb[wb.sheetnames[0]]
    if name in wb.sheetnames:
        return wb[name]
    # 容错：模板 sheet 名带首尾空格（如「合同 」「用途功能 」）时按去空格后唯一匹配
    hits = [s for s in wb.sheetnames if s.strip() == name]
    if len(hits) == 1:
        return wb[hits[0]]
    raise ValueError(f"模板中不存在工作表 '{name}'（现有：{', '.join(wb.sheetnames)}）")


def _resolve(path, ctx, row, where):
    try:
        return resolve(path, ctx, row)
    except ValueError as e:
        raise ValueError(f"{where} 取值失败（路径 '{path}'）：{e}") from e


def _write(ws, addr, value, where):
    try:
        ws[addr] = value
    except Exception as e:
        raise ValueError(f"{where} 写入失败：{e}") from e


def _fill_cells(ws, cells, ctx):
    for addr, path in (cells or {}).items():
        where = f"工作表 '{ws.title}' 单元格 {addr}"
        value = _resolve(path, ctx, None, where)
        _write(ws, addr, value, where)


def _fill_anchors(ws, anchors, ctx):
    for a in (anchors or []):
        find = a.get("find", "")
        col = (a.get("in_column") or "A").upper()
        offset_col = (a.get("offset_col") or col).upper()
        row_offset = int(a.get("row_offset") or 0)
        path = a.get("value", "")
        found = None
        for r in range(1, (ws.max_row or 1) + 1):
            v = ws[f"{col}{r}"].value
            if v is not None and find in str(v):
                found = r
                break
        if found is None:
            raise ValueError(
                f"工作表 '{ws.title}' 的 {col} 列未找到锚点文本 '{find}'")
        target = f"{offset_col}{found + row_offset}"
        where = f"工作表 '{ws.title}' 锚点 '{find}' → 单元格 {target}"
        value = _resolve(path, ctx, None, where)
        _write(ws, target, value, where)


# 公式里的单元格引用/区间：可带 sheet 前缀（'名 称'! 或 名称!），如
# =SUM(C8:C9)、=箱单!F10、='合同 '!B2。捕获组：sheet 前缀、起点、终点(可选)。
_REF_RE = re.compile(
    r"((?:'[^']+'|[A-Za-z0-9_一-龥]+)!)?"
    r"(\$?[A-Za-z]{1,3}\$?\d+)(?:(:)(\$?[A-Za-z]{1,3}\$?\d+))?")
_CELL_RE = re.compile(r"^(\$?[A-Za-z]{1,3})(\$?)(\d+)$")


def _shift_cell(ref, del_start, n, is_range_end=False, is_range_start=False):
    """单元格引用按删行调整。返回新引用或 None（=#REF!）。"""
    m = _CELL_RE.match(ref)
    if m is None:
        return ref
    col, dollar, row = m.group(1), m.group(2), int(m.group(3))
    if row >= del_start + n:
        row -= n
    elif row >= del_start:                      # 落在被删区间内
        if is_range_end:
            row = del_start - 1                 # 区间终点收缩到删除区上一行
        elif is_range_start:
            row = del_start                     # 区间起点贴到删除区起点（行号已让位）
        else:
            return None                         # 单引用 → #REF!
    return f"{col}{dollar}{row}"


def _adjust_formulas(wb, sheet_title, del_start, n):
    """sheet_title 中删除了 [del_start, del_start+n) 行后，调整全工作簿公式。"""
    quoted = f"'{sheet_title}'!"
    plain = f"{sheet_title}!"
    for ws in wb.worksheets:
        same_sheet = (ws.title == sheet_title)
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue

                def _rep(m):
                    prefix = m.group(1) or ""
                    # 该引用是否指向被删行的 sheet
                    if prefix:
                        target = prefix in (quoted, plain)
                    else:
                        target = same_sheet
                    if not target:
                        return m.group(0)
                    a, colon, b = m.group(2), m.group(3), m.group(4)
                    if colon:                       # 区间
                        na = _shift_cell(a, del_start, n, is_range_start=True)
                        nb = _shift_cell(b, del_start, n, is_range_end=True)
                        # 区间整体落入被删区 → #REF!
                        ra = int(_CELL_RE.match(na).group(3)) if na else None
                        rb = int(_CELL_RE.match(nb).group(3)) if nb else None
                        if ra is None or rb is None or ra > rb:
                            return "#REF!"
                        return f"{prefix}{na}:{nb}"
                    na = _shift_cell(a, del_start, n)
                    return f"{prefix}{na}" if na else "#REF!"

                new_v = _REF_RE.sub(_rep, v)
                if new_v != v:
                    cell.value = new_v


def _adjust_formulas_insert(wb, sheet_title, ins_start, n):
    """sheet_title 中在 ins_start 处插入了 n 行后，按 Excel 语义调整全工作簿公式。

    - 单引用：行号 >= ins_start → +n；
    - 区间：终点行 >= ins_start → 终点 +n；起点行 >= ins_start → 起点 +n
      （区间跨越插入点时只动终点 = Excel 的"区间扩张"，如 SUM(G17:G18)
      在 18 处插 17 行 → SUM(G17:G35)）。
    """
    quoted = f"'{sheet_title}'!"
    plain = f"{sheet_title}!"

    def _shift(ref):
        m = _CELL_RE.match(ref)
        if m is None:
            return ref
        col, dollar, row = m.group(1), m.group(2), int(m.group(3))
        if row >= ins_start:
            row += n
        return f"{col}{dollar}{row}"

    for ws in wb.worksheets:
        same_sheet = (ws.title == sheet_title)
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue

                def _rep(m):
                    prefix = m.group(1) or ""
                    if prefix:
                        target = prefix in (quoted, plain)
                    else:
                        target = same_sheet
                    if not target:
                        return m.group(0)
                    a, colon, b = m.group(2), m.group(3), m.group(4)
                    if colon:
                        return f"{prefix}{_shift(a)}:{_shift(b)}"
                    return f"{prefix}{_shift(a)}"

                new_v = _REF_RE.sub(_rep, v)
                if new_v != v:
                    cell.value = new_v


def _insert_rows_keep_merges(ws, ins_start, n):
    """在 ins_start 处插入 n 行（openpyxl insert_rows 不平移合并区，这里补齐）。

    - 整体在插入点下方的合并区：整体下移 n；
    - 跨越插入点的合并区：终点下移 n（区域扩张）；
    - 上方的不动。
    """
    plans = []
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row < ins_start:
            continue
        ws.unmerge_cells(str(rng))
        new_min = rng.min_row + n if rng.min_row >= ins_start else rng.min_row
        new_max = rng.max_row + n
        plans.append((rng.min_col, new_min, rng.max_col, new_max))
    ws.insert_rows(ins_start, n)
    for min_c, min_r, max_c, max_r in plans:
        if (min_r, min_c) != (max_r, max_c):
            ws.merge_cells(start_row=min_r, start_column=min_c,
                           end_row=max_r, end_column=max_c)


_ROWNUM_RE = re.compile(r"(\$?[A-Za-z]{1,3})(\$)?(\d+)")


def _shift_formula_relative(formula, offset):
    """模板块复制到下方块时，公式相对引用整体下移 offset 行（带 $ 的行不动）。
    复刻 Excel 拷贝粘贴语义（如块内 '=I20*G20' 复制到 +3 行 → '=I23*G23'）。"""
    def _rep(m):
        if m.group(2):                       # $ 锁行
            return m.group(0)
        return f"{m.group(1)}{int(m.group(3)) + offset}"
    return _ROWNUM_RE.sub(_rep, formula)


def _replicate_block(ws, start_row, row_step, count):
    """把模板块 [start_row, start_row+row_step) 的内容/样式/合并/行高
    复制到其后的 count-1 个块（行已预先插入）。值与公式一并复制，
    公式按相对引用平移——复刻 RPA「拷贝区域→粘贴」的行为（块内预填的
    固定列如 原产国/单位/币制/金额公式 随块复制，无须在映射里逐一登记）。"""
    max_col = ws.max_column or 1
    # 模板块内部的合并区（完全落在块内的）
    block_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row >= start_row and rng.max_row < start_row + row_step
    ]
    for i in range(1, count):
        offset = i * row_step
        for j in range(row_step):
            src_row = start_row + j
            dst_row = src_row + offset
            for c in range(1, max_col + 1):
                src = ws.cell(row=src_row, column=c)
                dst = ws.cell(row=dst_row, column=c)
                v = src.value
                if isinstance(v, str) and v.startswith("="):
                    v = _shift_formula_relative(v, offset)
                dst.value = v
                if src.has_style:
                    dst._style = copy(src._style)
            if src_row in ws.row_dimensions:
                ws.row_dimensions[dst_row].height = \
                    ws.row_dimensions[src_row].height
        for rng in block_merges:
            ws.merge_cells(start_row=rng.min_row + offset,
                           start_column=rng.min_col,
                           end_row=rng.max_row + offset,
                           end_column=rng.max_col)


def _delete_rows_keep_merges(ws, del_start, n):
    """删除 [del_start, del_start+n) 行。

    openpyxl 的 delete_rows 不处理合并单元格：合并区左上角的值会丢、
    区域定义不平移。这里先解除受影响的合并、删行、再按平移后的坐标
    重新合并（完全落在被删区间内的合并直接丢弃，跨界的收缩）。
    """
    plans = []          # (新区域字符串 或 None=丢弃)
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row < del_start:
            continue                            # 整体在删除区上方，不动
        ws.unmerge_cells(str(rng))
        min_r, max_r = rng.min_row, rng.max_row
        new_min = min_r - n if min_r >= del_start + n else (
            min_r if min_r < del_start else del_start)
        new_max = max_r - n if max_r >= del_start + n else (
            max_r if max_r < del_start else del_start - 1)
        if new_max >= new_min:
            plans.append((rng.min_col, new_min, rng.max_col, new_max))
    ws.delete_rows(del_start, n)
    for min_c, min_r, max_c, max_r in plans:
        if max_r >= min_r and (min_r, min_c) != (max_r, max_c):
            ws.merge_cells(start_row=min_r, start_column=min_c,
                           end_row=max_r, end_column=max_c)


def _copy_row_style(ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst._style = copy(src._style)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _fill_table(wb, ws, table, ctx, rows_source):
    source = table.get("source", "items")
    rows = None
    if rows_source and source in rows_source:
        rows = rows_source[source]
    elif source in ctx:
        rows = ctx[source]
    if rows is None:
        raise ValueError(
            f"工作表 '{ws.title}' 明细区数据源 '{source}' 不存在"
            "（可用：items / items_agg / plan_items / boxes / shipments）")
    start_row = int(table.get("start_row") or 1)
    # row_step：每条记录占几行（默认 1）。报关单一条明细占 3 行：
    #   D20 品名 / D21 申报要素 / G21 单位 / I22 币制 → row_step=3，
    #   columns 写成 {"D": 品名, "D+1": 申报要素, "G+1": 单位, "I+2": 币制}。
    row_step = max(1, int(table.get("row_step") or 1))
    # insert_rows 默认 False：货代/报关模板的明细区通常自带预留空行，且下方还有
    # 固定布局（落款、申报要素区、签章栏等固定行）；插行会把下方布局整体推乱。
    # 只有明细区真正开放式（下方无固定内容）的模板才显式配 "insert_rows": true。
    insert = bool(table.get("insert_rows"))
    columns = table.get("columns") or {}

    if insert and len(rows) > 1:
        # 复刻 RPA 行为：在首块之后插入 (n-1)*step 行，再把首块整体复制过去。
        # 模板若预留多于 1 个块（如采购合同预留 2 行），多余预留块顺势下移、
        # 保留为尾部空格式行——与历史成品逐格一致。
        n_extra = (len(rows) - 1) * row_step
        ins_at = start_row + row_step
        _insert_rows_keep_merges(ws, ins_at, n_extra)
        _adjust_formulas_insert(wb, ws.title, ins_at, n_extra)
        _replicate_block(ws, start_row, row_step, len(rows))

    # reserved_rows：模板预留 N 条记录的格式行；删行在写入前做，写入坐标按删后布局算。
    reserved = int(table.get("reserved_rows") or 0)
    if insert and reserved > 1:
        # 插行模板：首块后插了 (n-1) 块，模板第 2..N 个预留块整体下移成了空行段
        # （夹在明细与合计之间）——全部删掉，让合计/表尾紧跟明细（与 RPA 成品一致；
        # 报关箱单曾因此出现 12 行空档、合计被顶到 32 行）。
        del_start = start_row + len(rows) * row_step
        n_del = (reserved - 1) * row_step
        _delete_rows_keep_merges(ws, del_start, n_del)
        _adjust_formulas(wb, ws.title, del_start, n_del)
    elif reserved and len(rows) < reserved:
        # 非插行模板：记录少于预留时删掉多余预留行并调整公式
        del_start = start_row + len(rows) * row_step
        n_del = (reserved - len(rows)) * row_step
        _delete_rows_keep_merges(ws, del_start, n_del)
        _adjust_formulas(wb, ws.title, del_start, n_del)

    for i, row in enumerate(rows):
        base = start_row + i * row_step
        for col, path in columns.items():
            m = _COL_RE.match(str(col).strip())
            if m is None:
                raise ValueError(
                    f"工作表 '{ws.title}' 明细区({source}) 列名 '{col}' 不合法"
                    "（应为 'D' 或 'D+1' 形式）")
            letter, row_off = m.group(1).upper(), int(m.group(2) or 0)
            addr = f"{letter}{base + row_off}"
            where = f"工作表 '{ws.title}' 明细区({source}) 单元格 {addr}"
            value = _resolve(path, ctx, row, where)
            _write(ws, addr, value, where)


def _table_rows(table, ctx, rows_source):
    """取明细区数据源（fill / preview 共用）。不存在抛 ValueError。"""
    source = table.get("source", "items")
    if rows_source and source in rows_source:
        return source, rows_source[source]
    if source in ctx:
        return source, ctx[source]
    raise ValueError(
        f"明细区数据源 '{source}' 不存在"
        "（可用：items / items_agg / plan_items / boxes / shipments）")


def preview(template_path, mapping, ctx, rows_source=None, table_rows_limit=5):
    """同 fill 的取值/定位逻辑，但**不写文件**：返回将写入的清单。

    返回 {"cells": [{sheet, target, path, value[, anchor]}],
          "tables": [{sheet, source, start_row, row_step, total,
                      columns: [{col, path}], rows: [[{target, col, path, value}]]}],
          "errors": [str]}
    明细区只取前 table_rows_limit 行（total 给出总行数）。
    单项取值失败记入 errors，不中断其余项（预览容错）。
    """
    if not os.path.exists(template_path):
        raise ValueError(f"模板文件不存在：{template_path}")
    if not isinstance(mapping, dict) or not mapping.get("sheets"):
        raise ValueError("映射配置为空或缺少 sheets 节点，请先在模板上配置映射")

    out = {"cells": [], "tables": [], "errors": []}
    wb = load_workbook(template_path)
    for spec in mapping["sheets"]:
        try:
            ws = _get_sheet(wb, spec)
        except ValueError as e:
            out["errors"].append(str(e))
            continue
        title = ws.title

        # ---- 坐标单元格
        for addr, path in (spec.get("cells") or {}).items():
            try:
                v = _resolve(path, ctx, None, f"工作表 '{title}' 单元格 {addr}")
                out["cells"].append({"sheet": title, "target": addr,
                                     "path": path, "value": v})
            except ValueError as e:
                out["errors"].append(str(e))

        # ---- 锚点（在模板里定位目标格，逻辑与 _fill_anchors 一致）
        for a in (spec.get("anchors") or []):
            find = a.get("find", "")
            col = (a.get("in_column") or "A").upper()
            offset_col = (a.get("offset_col") or col).upper()
            row_offset = int(a.get("row_offset") or 0)
            path = a.get("value", "")
            found = None
            for r in range(1, (ws.max_row or 1) + 1):
                cv = ws[f"{col}{r}"].value
                if cv is not None and find in str(cv):
                    found = r
                    break
            if found is None:
                out["errors"].append(
                    f"工作表 '{title}' 的 {col} 列未找到锚点文本 '{find}'")
                continue
            target = f"{offset_col}{found + row_offset}"
            try:
                v = _resolve(path, ctx, None,
                             f"工作表 '{title}' 锚点 '{find}' → 单元格 {target}")
                out["cells"].append({"sheet": title, "target": target,
                                     "path": path, "value": v, "anchor": find})
            except ValueError as e:
                out["errors"].append(str(e))

        # ---- 明细区
        for table in (spec.get("tables") or []):
            try:
                source, rows = _table_rows(table, ctx, rows_source)
            except ValueError as e:
                out["errors"].append(f"工作表 '{title}' {e}")
                continue
            start_row = int(table.get("start_row") or 1)
            row_step = max(1, int(table.get("row_step") or 1))
            cols = []
            for col, path in (table.get("columns") or {}).items():
                m = _COL_RE.match(str(col).strip())
                if m is None:
                    out["errors"].append(
                        f"工作表 '{title}' 明细区({source}) 列名 '{col}' 不合法"
                        "（应为 'D' 或 'D+1' 形式）")
                    continue
                cols.append((str(col).strip(), m.group(1).upper(),
                             int(m.group(2) or 0), path))
            t_rows = []
            for i, row in enumerate(rows[:table_rows_limit]):
                base = start_row + i * row_step
                line = []
                for raw, letter, row_off, path in cols:
                    addr = f"{letter}{base + row_off}"
                    where = f"工作表 '{title}' 明细区({source}) 单元格 {addr}"
                    try:
                        v = _resolve(path, ctx, row, where)
                    except ValueError as e:
                        v = None
                        out["errors"].append(str(e))
                    line.append({"target": addr, "col": raw,
                                 "path": path, "value": v})
                t_rows.append(line)
            out["tables"].append({
                "sheet": title, "source": source,
                "start_row": start_row, "row_step": row_step,
                "total": len(rows),
                "columns": [{"col": c[0], "path": c[3]} for c in cols],
                "rows": t_rows,
            })
    return out


def fill(template_path, mapping, ctx, out_path, rows_source=None):
    """按映射把 ctx 填入模板，另存为 out_path。

    mapping: {"sheets": [{"name", "cells", "anchors", "tables"}, ...]}
    rows_source: 可选 {source名: 行列表}，覆盖 ctx 中同名数据源
                 （generate_service 用它过滤缺货代单号的货件等）。
    """
    if not os.path.exists(template_path):
        raise ValueError(f"模板文件不存在：{template_path}")
    if not isinstance(mapping, dict) or not mapping.get("sheets"):
        raise ValueError("映射配置为空或缺少 sheets 节点，请先在模板上配置映射")

    wb = load_workbook(template_path)
    # 顺序：cells → tables → anchors。anchors 放在 tables 之后，使锚点
    # 能定位到插行/删行之后的真实行（如箱单"合计"行、采购合同"付款方式"
    # 上方的交货时间行——它们的位置随明细行数浮动）。
    for spec in mapping["sheets"]:
        ws = _get_sheet(wb, spec)   # 显式确定的 ws 对象，后续写入只经由它
        _fill_cells(ws, spec.get("cells"), ctx)
        for table in (spec.get("tables") or []):
            _fill_table(wb, ws, table, ctx, rows_source)
        _fill_anchors(ws, spec.get("anchors"), ctx)

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    return out_path

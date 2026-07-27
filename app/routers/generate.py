"""生成路由：一键生成 / 文件下载删除 / 种子初始化（CONTRACTS.md）。"""

import os

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from .. import seed
from ..database import get_db
from ..models import GeneratedDoc
from ..services import contract_service, generate_service

router = APIRouter()


class GenerateIn(BaseModel):
    template_ids: list[int]


@router.post("/batches/{batch_id}/generate")
def generate(batch_id: int, body: GenerateIn, db: Session = Depends(get_db)):
    """强制先校验，passed 才生成；逐模板报错不互相影响。"""
    return generate_service.generate(db, batch_id, body.template_ids)


@router.post("/batches/{batch_id}/contracts")
def generate_contracts(batch_id: int, data: dict = None, db: Session = Depends(get_db)):
    """生成两级采购合同（朗格系两份：工厂→店铺、店铺→星盟；其它店一份，系数走 RuleConfig）。

    body: {persist: bool=true}——persist=false 只写文件不落 GeneratedDoc 记录（试跑用）。
    文件写入 output/{批次名-MMDD}/采购合同/，同名静默覆盖。
    返回 {generated:[绝对路径], warnings:[...]}。"""
    try:
        return contract_service.generate_two_level(
            db, batch_id, persist=bool((data or {}).get("persist", True)))
    except RuntimeError as e:
        db.rollback()
        raise HTTPException(404 if "不存在" in str(e) else 502, str(e))


@router.get("/batches/{batch_id}/preview")
def preview_generate(batch_id: int, template_id: int,
                     db: Session = Depends(get_db)):
    """生成前数据预览：返回将写入的清单（不写文件）。"""
    try:
        return generate_service.preview(db, batch_id, template_id)
    except ValueError as e:
        raise HTTPException(404 if "不存在" in str(e) else 400, str(e))


@router.get("/docs/{doc_id}/preview")
def preview_doc(doc_id: int, db: Session = Depends(get_db)):
    """已生成 Excel 文件预览：openpyxl data_only 读取，每个 sheet 返回
    非空区域（最多 60 行 × 20 列）的 {addr: value} 与合并区域列表。"""
    doc = db.get(GeneratedDoc, doc_id)
    if doc is None:
        raise HTTPException(404, f"生成记录 {doc_id} 不存在")
    if not doc.path or not os.path.exists(doc.path):
        raise HTTPException(404, f"文件已不存在：{doc.filename}")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(doc.path, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"文件读取失败：{e}")

    MAX_R, MAX_C = 60, 20
    sheets = []
    for ws in wb.worksheets:
        cap_r = min(ws.max_row or 1, MAX_R)
        cap_c = min(ws.max_column or 1, MAX_C)
        cells = {}
        used_r = used_c = 1
        for row in ws.iter_rows(min_row=1, max_row=cap_r,
                                min_col=1, max_col=cap_c):
            for cell in row:
                v = cell.value
                if v is None or v == "":
                    continue
                cells[cell.coordinate] = v
                used_r = max(used_r, cell.row)
                used_c = max(used_c, cell.column)
        merges = []
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= cap_r and rng.min_col <= cap_c:
                merges.append(str(rng))
                used_r = max(used_r, min(rng.max_row, cap_r))
                used_c = max(used_c, min(rng.max_col, cap_c))
        sheets.append({
            "name": ws.title,
            "rows": used_r, "cols": used_c,
            "cells": cells, "merges": merges,
            "truncated": (ws.max_row or 1) > MAX_R or (ws.max_column or 1) > MAX_C,
        })
    return {"doc_id": doc.id, "filename": doc.filename, "sheets": sheets}


@router.get("/docs/{doc_id}/download")
def download_doc(doc_id: int, db: Session = Depends(get_db)):
    doc = db.get(GeneratedDoc, doc_id)
    if doc is None:
        raise HTTPException(404, f"生成记录 {doc_id} 不存在")
    if not doc.path or not os.path.exists(doc.path):
        raise HTTPException(404, f"文件已不存在：{doc.filename}")
    return FileResponse(
        doc.path, filename=doc.filename,
        media_type="application/vnd.openxmlformats-officedocument"
                   ".spreadsheetml.sheet")


@router.delete("/docs/{doc_id}")
def delete_doc(doc_id: int, db: Session = Depends(get_db)):
    doc = db.get(GeneratedDoc, doc_id)
    if doc is None:
        raise HTTPException(404, f"生成记录 {doc_id} 不存在")
    if doc.path and os.path.exists(doc.path):
        try:
            os.remove(doc.path)
        except OSError:
            pass  # 文件占用等情况不阻塞记录删除
    db.delete(doc)
    db.commit()
    return {"ok": True}


@router.post("/seed/init")
def seed_init(db: Session = Depends(get_db)):
    """初始化：规则默认值种子 + 工厂信息库.xlsx 导入（import_service 就绪时）。"""
    return seed.run_seed(db)


@router.post("/seed/export-pkg")
def seed_export_pkg(db: Session = Depends(get_db)):
    """管理员导出本地初始化包 seed_pkg.zip（档案/规则/doc_rules/模板，离线分发勿进 git）。"""
    return seed.export_pkg(db)


@router.post("/seed/import-pkg")
async def seed_import_pkg(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """运营端导入初始化包（clone 后一次性执行；merge upsert 幂等可重导）。"""
    import tempfile
    data = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as f:
        f.write(data)
        tmp = f.name
    try:
        return seed.import_pkg(db, tmp)
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass

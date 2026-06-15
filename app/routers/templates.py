"""模板路由：上传 / AI 映射 / 试生成 / 字段字典（CONTRACTS.md，挂 /api 前缀由 main.py 做）。"""

import os
import re
import time

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from .. import ai_mapping
from ..database import TEMPLATE_STORE, get_db
from ..field_registry import FIELD_DICT
from ..models import Template
from ..services import generate_service

router = APIRouter()

_UNSAFE = re.compile(r'[\\/:*?"<>|\r\n\t]')

# 网格预览上限：与 docs preview 对齐，列扩到 A-Z（26 列）。
GRID_MAX_R, GRID_MAX_C = 60, 26


class AiMappingIn(BaseModel):
    stored_file: str


class PreviewGridIn(BaseModel):
    stored_file: str


def _read_grid(path: str):
    """openpyxl 读取 Excel，每个 sheet 返回非空区域（最多 60 行 × 26 列）的
    {addr: value}、合并区域列表、维度（与 generate.py /docs preview 结构对齐）。
    模板原件按文字读取（data_only=False），保留映射占位文字。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        cap_r = min(ws.max_row or 1, GRID_MAX_R)
        cap_c = min(ws.max_column or 1, GRID_MAX_C)
        cells = {}
        used_r = used_c = 1
        for row in ws.iter_rows(min_row=1, max_row=cap_r,
                                min_col=1, max_col=cap_c):
            for cell in row:
                v = cell.value
                if v is None or v == "":
                    continue
                cells[cell.coordinate] = v
                used_r = max(used_r, cell.row)
                used_c = max(used_c, cell.column)
        merges = []
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= cap_r and rng.min_col <= cap_c:
                merges.append(str(rng))
                used_r = max(used_r, min(rng.max_row, cap_r))
                used_c = max(used_c, min(rng.max_col, cap_c))
        sheets.append({
            "name": ws.title,
            "rows": used_r, "cols": used_c,
            "cells": cells, "merges": merges,
            "truncated": (ws.max_row or 1) > GRID_MAX_R
                         or (ws.max_column or 1) > GRID_MAX_C,
        })
    wb.close()
    return sheets


class TestGenerateIn(BaseModel):
    batch_id: int


@router.post("/templates/upload-file")
async def upload_file(file: UploadFile = File(...)):
    """存 templates_store/{时间戳}_{安全文件名}，返回 sheet 名列表。"""
    original = file.filename or "template.xlsx"
    safe = _UNSAFE.sub("_", os.path.basename(original)).strip() or "template.xlsx"
    if not safe.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "只支持 .xlsx / .xlsm 模板文件")
    stored = f"{time.strftime('%Y%m%d%H%M%S')}_{safe}"
    path = os.path.join(TEMPLATE_STORE, stored)
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    try:
        wb = load_workbook(path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
    except Exception as e:
        os.remove(path)
        raise HTTPException(400, f"无法读取 Excel 文件：{e}")
    return {"stored_file": stored, "original_name": original,
            "sheetnames": sheetnames}


@router.post("/templates/ai-mapping")
def ai_mapping_draft(body: AiMappingIn):
    """AI 映射草稿。未配置 API key / 调用失败时返回 {"error": ...}，HTTP 200。"""
    return ai_mapping.suggest_mapping(body.stored_file)


@router.post("/templates/{template_id}/test-generate")
def test_generate(template_id: int, body: TestGenerateIn,
                  db: Session = Depends(get_db)):
    """单模板试生成（不落 GeneratedDoc），直接返回文件下载。"""
    try:
        out_path, filename = generate_service.test_generate(
            db, template_id, body.batch_id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return FileResponse(
        out_path, filename=filename,
        media_type="application/vnd.openxmlformats-officedocument"
                   ".spreadsheetml.sheet")


@router.get("/templates/{template_id}/grid")
def template_grid(template_id: int, db: Session = Depends(get_db)):
    """模板原件网格预览（供前端可视化点选映射）：读 templates_store/{stored_file}，
    每 sheet 返回非空区域 {addr: value}、合并区域、维度。无文件返回 404。"""
    t = db.get(Template, template_id)
    if t is None:
        raise HTTPException(404, f"模板 {template_id} 不存在")
    if not t.stored_file:
        raise HTTPException(404, "该模板未关联原件")
    path = os.path.join(TEMPLATE_STORE, os.path.basename(t.stored_file))
    if not os.path.exists(path):
        raise HTTPException(404, f"模板原件已不存在：{t.stored_file}")
    try:
        sheets = _read_grid(path)
    except Exception as e:
        raise HTTPException(400, f"文件读取失败：{e}")
    return {"template_id": t.id, "name": t.name,
            "stored_file": t.stored_file, "sheets": sheets}


@router.post("/templates/preview-grid")
def template_preview_grid(body: PreviewGridIn):
    """按 stored_file 直接读网格（向导中刚上传、尚未存模板记录时预览）。"""
    safe = os.path.basename(body.stored_file or "")
    if not safe:
        raise HTTPException(400, "缺少 stored_file")
    path = os.path.join(TEMPLATE_STORE, safe)
    if not os.path.exists(path):
        raise HTTPException(404, f"文件不存在：{body.stored_file}")
    try:
        sheets = _read_grid(path)
    except Exception as e:
        raise HTTPException(400, f"文件读取失败：{e}")
    return {"stored_file": body.stored_file, "sheets": sheets}


@router.get("/fields")
def fields():
    """字段字典（前端映射配置面板用）。"""
    return FIELD_DICT

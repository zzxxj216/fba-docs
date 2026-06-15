"""AI 映射助手：读上传的 Excel 模板 → 调 Anthropic API → 映射草稿 + 置信度。

- 模型 claude-opus-4-8，adaptive thinking，structured outputs
  （output_config.format json_schema；动态键的 mapping 用 entries 数组返回，
  Python 侧组装成 DESIGN.md §6 的映射 JSON）。
- ANTHROPIC_API_KEY 未配置或调用失败 → 返回 {"error": "..."} 优雅降级，不抛异常。
"""

import json
import os

from openpyxl import load_workbook

from .database import TEMPLATE_STORE
from .field_registry import FIELD_DICT

MODEL = "claude-opus-4-8"
MAX_ROWS = 60  # 每个 sheet 提取前 60 行非空单元格

# structured outputs：对象一律 additionalProperties:false、字段全 required，
# 不适用的字段由模型填 ""/0。
_OUTPUT_SCHEMA = {
    "type": "object",
    "properties": {
        "entries": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string", "description": "工作表名"},
                    "kind": {"type": "string",
                             "enum": ["cell", "anchor", "table_column"]},
                    "address": {
                        "type": "string",
                        "description": "cell: 单元格如 B5；table_column: 列字母如 G；anchor: 留空",
                    },
                    "field": {"type": "string",
                              "description": "字段字典路径，如 shipment.fc_code；固定文字用 text: 前缀"},
                    "anchor_find": {"type": "string",
                                    "description": "kind=anchor 时：要查找的锚点文本，否则留空"},
                    "anchor_in_column": {"type": "string",
                                         "description": "kind=anchor 时：在哪一列找（列字母），否则留空"},
                    "anchor_offset_col": {"type": "string",
                                          "description": "kind=anchor 时：写入哪一列（列字母），否则留空"},
                    "anchor_row_offset": {"type": "integer",
                                          "description": "kind=anchor 时：行偏移，通常 0"},
                    "table_source": {"type": "string",
                                     "description": "kind=table_column 时：items/boxes/shipments，否则留空"},
                    "table_start_row": {"type": "integer",
                                        "description": "kind=table_column 时：明细起始行号，否则 0"},
                    "confidence": {"type": "number",
                                   "description": "0~1，不确定就给低值"},
                    "note": {"type": "string", "description": "判断依据，简短中文"},
                },
                "required": ["sheet", "kind", "address", "field",
                             "anchor_find", "anchor_in_column",
                             "anchor_offset_col", "anchor_row_offset",
                             "table_source", "table_start_row",
                             "confidence", "note"],
                "additionalProperties": False,
            },
        },
        "notes": {"type": "string", "description": "整体说明：模板结构、没把握的地方"},
    },
    "required": ["entries", "notes"],
    "additionalProperties": False,
}


def _extract_sheet_text(path):
    """每个 sheet 前 MAX_ROWS 行非空单元格清单（坐标=值）+ 合并区域。"""
    wb = load_workbook(path, data_only=False)
    parts = []
    for ws in wb.worksheets:
        lines = [f"### 工作表: {ws.title}"]
        merged = [str(r) for r in ws.merged_cells.ranges]
        if merged:
            lines.append("合并区域: " + ", ".join(merged))
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(MAX_ROWS, ws.max_row or 1)):
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                lines.append(f"{cell.coordinate}={cell.value}")
                count += 1
        if count == 0:
            lines.append("(空表)")
        parts.append("\n".join(lines))
    return "\n\n".join(parts), wb.sheetnames


def _field_dict_text():
    lines = []
    for ns, fields in FIELD_DICT.items():
        lines.append(f"[{ns}.*]")
        for key, label in fields.items():
            lines.append(f"  {ns}.{key} —— {label}")
    return "\n".join(lines)


_SYSTEM = """你是 FBA 发货文件系统的模板映射助手。用户上传了一个货代/报关 Excel 模板，\
你要分析模板结构，给出"模板单元格 ↔ 系统字段"的映射草稿。

## 可用字段字典（映射的 field 只能取这里的路径；固定文字用 "text:文字" 形式）
{fields}

## 映射的三种定位方式（输出 entries 数组，由系统组装成映射 JSON）
1. kind=cell：固定坐标。address 填要**写入**的单元格（通常是标签旁边/下面的空格，\
不是标签本身），field 填字段路径。
2. kind=anchor：锚点行定位（标签文本位置不固定时用）。anchor_find=标签文本、\
anchor_in_column=标签所在列、anchor_offset_col=写入列、anchor_row_offset=行偏移(通常0)。
3. kind=table_column：明细区按行重复填充。table_source=items(SKU明细)/boxes(逐箱)/\
shipments(每货件一行)，table_start_row=第一条数据行的行号（表头的下一行），\
address 填列字母，field 填行内字段（item.*/box.*/shipment.*，calc.* 也按行取值）。

## 要求
- 模板里已印好的标签文字不要映射；要找的是标签对应的**数据写入位置**。
- 同一明细区的多个列输出多条 table_column，table_source 和 table_start_row 保持一致。
- 日期、编号、汇总（总箱数/总毛重等）优先用 calc.* / meta.* / shipment.total_*。
- 不确定的也要给出，但 confidence 给低（<0.5），note 里写明疑点。
- notes 里总结模板结构与所有没把握的地方。"""


def suggest_mapping(stored_file):
    """→ {"mapping": {...}, "confidences": [...], "notes": "..."} 或 {"error": "..."}"""
    path = stored_file or ""
    if not os.path.isabs(path):
        path = os.path.join(TEMPLATE_STORE, path)
    if not os.path.exists(path):
        return {"error": f"模板文件不存在：{stored_file}"}

    try:
        sheet_text, _ = _extract_sheet_text(path)
    except Exception as e:
        return {"error": f"读取模板失败：{e}"}

    if not (os.getenv("ANTHROPIC_API_KEY") or "").strip():
        return {"error": "未配置 ANTHROPIC_API_KEY，请在 .env 中填写后重试"}

    try:
        import anthropic
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=MODEL,
            max_tokens=16000,
            thinking={"type": "adaptive"},
            system=_SYSTEM.format(fields=_field_dict_text()),
            messages=[{
                "role": "user",
                "content": ("以下是模板的非空单元格清单（坐标=值）和合并区域，"
                            "请给出映射草稿：\n\n" + sheet_text),
            }],
            output_config={"format": {"type": "json_schema",
                                      "schema": _OUTPUT_SCHEMA}},
        )
        if response.stop_reason == "refusal":
            return {"error": "AI 拒绝了本次请求，请检查模板内容后重试"}
        text = next((b.text for b in response.content if b.type == "text"), "")
        data = json.loads(text)
    except Exception as e:
        return {"error": f"AI 映射调用失败：{e}"}

    return _assemble(data)


def _assemble(data):
    """entries 数组 → DESIGN.md §6 映射 JSON + confidences。"""
    sheets = {}       # name -> {"name", "cells", "anchors", "tables"}
    tables_idx = {}   # (sheet, source, start_row) -> table dict
    confidences = []

    for e in data.get("entries", []):
        name = e.get("sheet") or ""
        sp = sheets.setdefault(name, {"name": name, "cells": {},
                                      "anchors": [], "tables": []})
        kind = e.get("kind")
        field = e.get("field") or ""
        addr = (e.get("address") or "").strip()
        target = f"{name}!{addr}" if addr else name

        if kind == "cell" and addr:
            sp["cells"][addr] = field
        elif kind == "anchor":
            sp["anchors"].append({
                "find": e.get("anchor_find") or "",
                "in_column": (e.get("anchor_in_column") or "A").upper(),
                "offset_col": (e.get("anchor_offset_col") or "A").upper(),
                "row_offset": int(e.get("anchor_row_offset") or 0),
                "value": field,
            })
            target = f"{name}!锚点[{e.get('anchor_find', '')}]"
        elif kind == "table_column" and addr:
            source = e.get("table_source") or "items"
            start_row = int(e.get("table_start_row") or 1)
            key = (name, source, start_row)
            if key not in tables_idx:
                tables_idx[key] = {"source": source, "start_row": start_row,
                                   "insert_rows": True, "columns": {}}
                sp["tables"].append(tables_idx[key])
            tables_idx[key]["columns"][addr.upper()] = field
            target = f"{name}!{addr.upper()}{start_row}+"
        else:
            continue

        confidences.append({
            "target": target,
            "field": field,
            "confidence": e.get("confidence", 0),
            "note": e.get("note") or "",
        })

    mapping = {"sheets": [sp for sp in sheets.values()
                          if sp["cells"] or sp["anchors"] or sp["tables"]]}
    return {"mapping": mapping, "confidences": confidences,
            "notes": data.get("notes", "")}

# -*- coding: utf-8 -*-
"""加测批次 HUHOLE-US-6.12（跨运通）：报关资料 + 采购合同 + 投保仓库货值。

- 投保单不生成：成品 D3/F3/G3（货代单号）为空——RPA 在未拿到货代单号时就出了
  投保文件；本系统 requires_forwarder_no=True 设计为强制回填后生成（P2 拍板），
  此差异在报告中说明。
- 6.12 托书为又一变体模板（sheet=发票模板/Sheet1），不在本轮范围，报告说明。
"""
import sys

sys.path.insert(0, r"D:\amazon")

from app.database import SessionLocal
from app.models import Batch, Forwarder
from app.services import generate_service, import_service

CASE = r"C:\Users\zane\Downloads\HUBFSE自动化"
HU12 = CASE + r"\HUHOLE-US-6.12"
db = SessionLocal()
FWD_K = db.query(Forwarder).filter(Forwarder.name.like("%跨运通%")).first().id

# 1) 产品刷新为 6.12 周口径（报关价=试算 6.12HU-US；品名/箱重=6.12 补仓计划）
from openpyxl import load_workbook

from app.models import Product


def import_plan_products(plan_path, price_path, price_sheet, declare):
    wb = load_workbook(plan_path, data_only=True, read_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr) if h}
    prices = {}
    wb = load_workbook(price_path, data_only=True, read_only=True)
    for r in wb[price_sheet].iter_rows(values_only=True):
        if r and len(r) > 6 and r[2] and r[6] is not None \
                and str(r[2]).strip() != "sku":
            try:
                prices[str(r[2]).strip()] = float(r[6])
            except (TypeError, ValueError):
                continue
    wb.close()

    def g(r, name):
        i = col.get(name)
        return r[i] if i is not None and i < len(r) else None

    n = 0
    for r in rows[1:]:
        sku = g(r, "AMAZON-SKU")
        if not sku:
            continue
        sku = str(sku).strip()
        p = db.query(Product).filter(Product.sku == sku).first()
        if p is None:
            p = Product(sku=sku)
            db.add(p)
        p.name_customs_cn = str(g(r, "中文品名") or "").strip()
        p.name_customs_en = str(g(r, "英文品名") or "").strip()
        hs = g(r, "海关编码")
        if hs:
            p.hs_code = str(hs).strip()
        p.material = str(g(r, "材质") or "").strip()
        p.usage = str(g(r, "用途") or "").strip()
        p.brand_name = str(g(r, "品牌") or "").strip()
        if g(r, "每箱重量（kg )") is not None:
            p.box_weight_kg = float(g(r, "每箱重量（kg )"))
        if g(r, "每箱商品数") is not None:
            p.qty_per_box = int(g(r, "每箱商品数"))
        if g(r, "单品单价(¥)") is not None:
            p.purchase_cost_default = float(g(r, "单品单价(¥)"))
        if sku in prices:
            p.unit_price_default = prices[sku]
        if declare:
            p.declare_elements = declare
        n += 1
        p.sort_index = n          # 当周计划行序（采购合同聚合行序）
        db.flush()
    db.commit()
    return n


n = import_plan_products(
    HU12 + r"\补仓计划-US-6.12-跨运通.xlsx",
    HU12 + r"\HU2026报关价格试算.xlsx", "6.12HU-US",
    "用途：车库挂钩|材质：铁|品牌：HUHOLE")
print("6.12 products:", n)

# 2) 导入批次
r = import_service.import_shipping_order(
    db, HU12 + r"\HUHOLE-FBA货件-922523477085970432.xlsx",
    filename="HUHOLE-FBA货件-922523477085970432.xlsx",
    batch_name="HUHOLE-US-0612K")
print("batch:", r["batch_id"], "passed:", r["report"]["passed"])
for e in r["report"]["errors"]:
    print("  ", e["level"], e["msg"][:90])
b = db.get(Batch, r["batch_id"])
b.base_date = "2026-06-12"
b.contract_date = "2026-05-12"
for sp in b.shipments:
    sp.forwarder_id = FWD_K          # 货代单号成品期未签发，投保单按设计跳过
db.commit()

# 3) 生成：报关 + 采购合同 + 投保仓库货值
from app.models import Template
ids = [db.query(Template).filter(Template.name == n2).first().id
       for n2 in ("HUHOLE-报关资料", "HUHOLE-采购合同", "HUBFSE-投保仓库货值")]
res = generate_service.generate(db, b.id, ids)
if res.get("blocked"):
    print("BLOCKED")
    for e in res["report"]["errors"]:
        print("  ", e["level"], e["msg"])
else:
    print("generated:", len(res["generated"]))
    for g in res["generated"]:
        print("  ", g["filename"])
    for e in res["errors"]:
        print("  ERR:", e)
db.close()
print("done")
